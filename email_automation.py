import html
import functools
import json
import logging
import mimetypes
import os
import random
import re
import shutil
import smtplib
import subprocess
import sys
import threading
import time
import uuid
from urllib.parse import quote
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from dotenv import dotenv_values
from openpyxl import Workbook, load_workbook
from account_service import AccountService
from attachment_library_service import AttachmentLibraryService
from bounce_tracking_service import BounceTrackingService
from build_info import BUILD_TIMESTAMP_UTC
from placeholder_service import PlaceholderService
from reply_tracking_service import ReplyTrackingService
from tracking_sync_service import TrackingSynchronizationService, synchronization_is_due


APP_NAME = "Email Automation"
TASK_NAME = "The Power People Daily Email Automation"
DEFAULT_DATA_DIR = r"F:\CODEX\Email_automation"
DEFAULT_CONFIG = {"daily_limit": 5, "schedule_time": "09:00", "data_dir": DEFAULT_DATA_DIR, "default_sender_name": "The Power People", "random_delay_min": 5, "random_delay_max": 15, "retry_count": 1, "backup_enabled": True, "theme": "Light", "tracking_sync_enabled": False, "tracking_sync_interval_hours": 5, "tracking_last_sync_time": "", "FullSynchronizationDebug": True}
TRACKING_BASE_URL = "https://emailtrackingserver-v2-2.onrender.com"
SEND_REGISTRATION_ENDPOINT = "/api/tracking/register-send"
PAGE_SIZE = 20


def exe_dir():
    return Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent


def project_dir():
    current = Path(sys.executable if getattr(sys, "frozen", False) else __file__).resolve()
    for parent in [current.parent, *current.parents]:
        if parent.name == "EmailAutomation" and parent.parent.name.lower() == "email automation v2":
            return parent
    return Path(__file__).resolve().parent


def local_appdata_dir():
    root = project_dir()
    if root.name == "EmailAutomation" and root.parent.name.lower() == "email automation v2":
        return root / ".runtime" / "LocalAppData"
    return Path(os.environ.get("LOCALAPPDATA", str(Path.home())))


def install_dir():
    return local_appdata_dir() / "ThePowerPeople" / "EmailAutomation"


def config_path():
    return install_dir() / "config" / "settings.json"


def load_config():
    p = config_path()
    if not p.exists():
        return dict(DEFAULT_CONFIG)
    try:
        return {**DEFAULT_CONFIG, **json.loads(p.read_text(encoding="utf-8"))}
    except Exception:
        return dict(DEFAULT_CONFIG)


def save_config(cfg):
    config_path().parent.mkdir(parents=True, exist_ok=True)
    config_path().write_text(json.dumps(cfg, indent=2), encoding="utf-8")


def app_paths(cfg=None):
    cfg = cfg or load_config()
    data = Path(cfg["data_dir"])
    root = install_dir()
    return {
        "root": root,
        "data": data,
        "list": data / "mail_list.xlsx",
        "env": root / ".env",
        "backup": root / "backup",
        "debug": root / "debug",
        "accounts": root / "config" / "accounts.json",
        "schedules": root / "config" / "schedules.json",
    }


def ensure_dirs():
    p = app_paths()
    for key in ("root", "backup"):
        p[key].mkdir(parents=True, exist_ok=True)


def tracking_registration_debug_log(**values):
    try:
        path = app_paths()["debug"] / "tracking-registration-debug.log"
        path.parent.mkdir(parents=True, exist_ok=True)
        entry = {"timestamp": datetime.now(timezone.utc).isoformat(), **values}
        with path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(entry, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass


class TrackingRegistrationFileHandler(logging.Handler):
    def __init__(self, path):
        super().__init__(logging.DEBUG)
        self.path = Path(path)

    def emit(self, record):
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            message = self.format(record)
            with self.path.open("a", encoding="utf-8") as handle:
                handle.write(message + "\n")
        except Exception:
            self.handleError(record)


def tracking_registration_logger():
    logger = logging.getLogger("email_automation.tracking_registration")
    logger.setLevel(logging.DEBUG)
    logger.propagate = False
    path = app_paths()["debug"] / "tracking-registration-debug.log"
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = False
    for handler in list(logger.handlers):
        if isinstance(handler, TrackingRegistrationFileHandler) and handler.path == path:
            existing = True
        else:
            logger.removeHandler(handler)
            try:
                handler.close()
            except Exception:
                pass
    if not existing:
        handler = TrackingRegistrationFileHandler(path)
        handler.setLevel(logging.DEBUG)
        handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        logger.addHandler(handler)
    return logger


def registration_failure(reason, **values):
    tracking_registration_debug_log(event="Registration Failed", reason=reason, **values)
    print(f"Email tracking registration failed: {reason}")


def generate_message_id():
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S%f")
    return f"<{uuid.uuid4().hex}.{timestamp}@emailautomation-v2.local>"


def register_sent_email(tracking_id, sender_mail, recipient_mail, mail_subject, project_name, excel_file_path, sent_time, message_id):
    logger = tracking_registration_logger()
    payload = {
        "tracking_id": str(tracking_id or ""),
        "message_id": str(message_id or ""),
        "sender_mail": str(sender_mail or ""),
        "recipient_mail": str(recipient_mail or ""),
        "mail_subject": str(mail_subject or ""),
        "project_name": str(project_name or ""),
        "excel_file_path": str(excel_file_path or ""),
        "sent_time": str(sent_time or ""),
    }
    base_url = str(TRACKING_BASE_URL or "").strip().rstrip("/")
    missing = [name for name in ("tracking_id", "message_id", "recipient_mail", "sender_mail") if not payload[name]]
    if not base_url:
        missing.append("TRACKING_BASE_URL")
    url = f"{base_url}{SEND_REGISTRATION_ENDPOINT}" if base_url else SEND_REGISTRATION_ENDPOINT
    logger.info("Calling POST /api/tracking/register-send")
    logger.info("Full request URL: %s", url)
    logger.info("JSON payload: %s", json.dumps(payload, ensure_ascii=False, default=str))
    if missing:
        registration_failure(
            "Missing required tracking registration value(s): " + ", ".join(missing),
            request_url=url,
            http_method="POST",
            payload=payload,
        )
        return False
    data = json.dumps(payload).encode("utf-8")
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": "EmailAutomation/2",
    }
    for attempt in (1, 2):
        request = Request(url, data=data, method="POST", headers=headers)
        tracking_registration_debug_log(event="Registration Request", request_url=url, http_method="POST", payload=payload, attempt=attempt)
        try:
            with urlopen(request, timeout=20) as response:
                body = response.read().decode("utf-8", errors="replace")
                status_code = getattr(response, "status", None) or (response.getcode() if hasattr(response, "getcode") else 200)
            logger.info("HTTP status code: %s", status_code)
            logger.info("Response body: %s", body)
            tracking_registration_debug_log(
                event="Registration Response",
                request_url=url,
                http_method="POST",
                payload=payload,
                response_status_code=status_code,
                response_body=body,
                attempt=attempt,
            )
            if 200 <= int(status_code) < 300:
                return True
            registration_failure(
                f"Tracking registration API returned HTTP {status_code}",
                request_url=url,
                http_method="POST",
                payload=payload,
                response_status_code=status_code,
                response_body=body,
                attempt=attempt,
            )
            return False
        except HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace") if hasattr(exc, "read") else ""
            logger.info("HTTP status code: %s", exc.code)
            logger.info("Response body: %s", body)
            logger.exception("Any exception")
            registration_failure(
                f"Tracking registration API returned HTTP {exc.code}",
                request_url=url,
                http_method="POST",
                payload=payload,
                response_status_code=exc.code,
                response_body=body,
                exception=str(exc),
                attempt=attempt,
            )
            return False
        except (URLError, TimeoutError, OSError) as exc:
            logger.exception("Any exception")
            tracking_registration_debug_log(
                event="Registration Network Error",
                request_url=url,
                http_method="POST",
                payload=payload,
                exception=repr(exc),
                attempt=attempt,
                will_retry=attempt == 1,
            )
            if attempt == 1:
                time.sleep(2)
                continue
            registration_failure(
                f"Tracking registration network error after retry: {exc}",
                request_url=url,
                http_method="POST",
                payload=payload,
                exception=repr(exc),
                attempt=attempt,
            )
            return False
        except Exception as exc:
            logger.exception("Any exception")
            registration_failure(
                f"Tracking registration failed: {exc}",
                request_url=url,
                http_method="POST",
                payload=payload,
                exception=repr(exc),
                attempt=attempt,
            )
            return False
    return False


def account_service():
    return AccountService(app_paths()["accounts"])


def tracking_sync_service():
    return TrackingSynchronizationService(TRACKING_BASE_URL, None, None)


def bounce_tracking_service():
    paths = app_paths()
    return BounceTrackingService(TRACKING_BASE_URL, paths["debug"] / "processed-bounces.json", paths["debug"])


def reply_tracking_service():
    return ReplyTrackingService(TRACKING_BASE_URL, app_paths()["debug"])


def run_bounce_detection():
    accounts = account_service()
    service = bounce_tracking_service()
    totals = {"detected": 0, "registered": 0, "duplicates": 0, "skipped": 0, "errors": 0}
    for account in accounts.list_accounts():
        configuration = accounts.imap_configuration(account["email"]) if account.get("enabled", True) else None
        if not configuration:
            continue
        try:
            result = service.check_account(configuration)
            for key in totals:
                totals[key] += int(result.get(key, 0))
        except Exception as exc:
            totals["errors"] += 1
            service._log("Errors", email=account.get("email"), error=str(exc))
    return totals


def synchronize_with_bounce(last_sync_time="", full_synchronization_debug=False):
    first = tracking_sync_service().sync(last_sync_time, full_synchronization_debug)
    bounce = run_bounce_detection()
    final = dict(first)
    final["bounce_detected"] = bounce["detected"]
    final["bounce_registered"] = bounce["registered"]
    final["bounce_duplicates"] = bounce["duplicates"]
    final["bounce_skipped"] = bounce["skipped"]
    final["bounce_errors"] = bounce["errors"]
    final["post_bounce_sync_ran"] = False
    if bounce["registered"]:
        second = tracking_sync_service().sync(first.get("last_sync_time", last_sync_time), full_synchronization_debug)
        final["post_bounce_sync_ran"] = True
        final["records_downloaded"] = first.get("records_downloaded", 0) + second.get("records_downloaded", 0)
        final["rows_updated"] = first.get("rows_updated", 0) + second.get("rows_updated", 0)
        final["download_records_received"] = first.get("download_records_received", 0) + second.get("download_records_received", 0)
        final["download_rows_updated"] = first.get("download_rows_updated", 0) + second.get("download_rows_updated", 0)
        final["bounce_records_received"] = first.get("bounce_records_received", 0) + second.get("bounce_records_received", 0)
        final["bounce_rows_updated"] = first.get("bounce_rows_updated", 0) + second.get("bounce_rows_updated", 0)
        final["execution_time"] = first.get("execution_time", 0) + second.get("execution_time", 0)
        final["last_sync_time"] = second.get("last_sync_time", first.get("last_sync_time", last_sync_time))
    return final


def attachment_library_service():
    return AttachmentLibraryService(TRACKING_BASE_URL, None)


def resolve_selected_attachments(attachment_ids):
    selected_ids = [str(value) for value in (attachment_ids or [])]
    if not selected_ids:
        return []
    available = {item["id"]: item for item in attachment_library_service().list_attachments()}
    missing = [value for value in selected_ids if value not in available]
    if missing:
        raise ValueError("One or more selected attachments have been deleted from the server. Refresh the attachment list and try again.")
    return [available[value] for value in selected_ids]


def build_attachment_links_html(tracking_id, attachments):
    if not attachments:
        return "", []
    rows = ["<p><strong>Attachments</strong></p>"]
    urls = []
    for attachment in attachments:
        url = f"{TRACKING_BASE_URL}/download/{tracking_id}/{quote(str(attachment['id']), safe='')}"
        urls.append(url)
        rows.append(f'<p>📄 <a href="{html.escape(url, quote=True)}">{html.escape(attachment["file_name"])}</a></p>')
    return "".join(rows), urls


def register_attachment_mapping(tracking_id, attachments):
    if not attachments:
        return
    attachment_library_service().register_tracking_attachments(tracking_id, [item["id"] for item in attachments])


def local_attachment_info(path):
    file_path = Path(path)
    size = file_path.stat().st_size if file_path.exists() else 0
    return {"path": str(file_path), "name": file_path.name, "size": size}


def format_file_size(size):
    value = float(size or 0)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.0f} {unit}" if unit == "B" else f"{value:.1f} {unit}"
        value /= 1024
    return f"{size} B"


def attach_local_file(message, path):
    file_path = Path(path)
    content_type, _encoding = mimetypes.guess_type(str(file_path))
    if not content_type:
        content_type = "application/octet-stream"
    maintype, subtype = content_type.split("/", 1)
    with file_path.open("rb") as handle:
        message.add_attachment(handle.read(), maintype=maintype, subtype=subtype, filename=file_path.name)


def local_attachment_warning(message, **values):
    tracking_registration_debug_log(event="Local Attachment Warning", warning=message, **values)
    print(f"Local attachment warning: {message}")


def migrate_legacy_account():
    service = account_service()
    env_path = app_paths()["env"]
    if service.list_accounts():
        if env_path.exists():
            try: env_path.unlink()
            except Exception: pass
        return
    try:
        address, password = credentials()
        service.save_account("Primary Gmail", address, password, True)
        env_path.unlink(missing_ok=True)
    except Exception:
        pass


def credentials():
    values = dotenv_values(app_paths()["env"])
    address = (values.get("EMAIL_ADDRESS") or "").strip()
    password = (values.get("EMAIL_PASSWORD") or "").strip().replace(" ", "")
    if not address or not password:
        raise ValueError("EMAIL_ADDRESS or EMAIL_PASSWORD is missing from .env")
    return address, password


def header_map(ws):
    return {str(cell.value).strip().lower(): i for i, cell in enumerate(ws[1], 1) if cell.value is not None}


@functools.lru_cache(maxsize=2048)
def domain_has_mx(domain):
    """Return True only when the recipient domain publishes a usable MX record."""
    try:
        result = subprocess.run(
            ["nslookup", "-type=MX", domain],
            capture_output=True,
            text=True,
            timeout=15,
            startupinfo=hidden_startupinfo(),
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    except Exception:
        return False
    output = f"{result.stdout}\n{result.stderr}"
    exchangers = re.findall(r"mail exchanger\s*=\s*([^\s]+)", output, flags=re.IGNORECASE)
    return any(value.rstrip(".") for value in exchangers)


def validate_recipient_email(email):
    """Validate common email syntax, DNS-safe domain syntax, and MX availability."""
    value = str(email or "").strip()
    if len(value) > 254 or value.count("@") != 1:
        return False
    local, domain = value.rsplit("@", 1)
    if not local or len(local) > 64 or not domain:
        return False
    if not re.fullmatch(r"[A-Za-z0-9!#$%&'*+/=?^_`{|}~.-]+", local):
        return False
    if local.startswith(".") or local.endswith(".") or ".." in local:
        return False
    try:
        ascii_domain = domain.rstrip(".").encode("idna").decode("ascii").lower()
    except UnicodeError:
        return False
    if not ascii_domain or len(ascii_domain) > 253 or "." not in ascii_domain:
        return False
    labels = ascii_domain.split(".")
    if any(not label or len(label) > 63 or label.startswith("-") or label.endswith("-") or not re.fullmatch(r"[a-z0-9-]+", label) for label in labels):
        return False
    return domain_has_mx(ascii_domain)


def validate_sheet(ws):
    h = header_map(ws)
    missing = [name for name in ("email", "subject", "body", "status") if name not in h]
    if missing:
        raise ValueError("Missing Excel columns: " + ", ".join(x.title() for x in missing))
    if "sentdate" not in h:
        col = ws.max_column + 1
        ws.cell(1, col, "SentDate")
        h["sentdate"] = col
    if "result" not in h:
        col = ws.max_column + 1
        ws.cell(1, col, "Result")
        h["result"] = col
    return h


def ensure_tracking_column(workbook, worksheet, workbook_path):
    """Return the TrackingId column, creating and immediately saving it if absent."""
    headers = header_map(worksheet)
    if "trackingid" in headers:
        return headers["trackingid"]
    column = worksheet.max_column + 1
    worksheet.cell(1, column, "TrackingId")
    workbook.save(workbook_path)
    return column


def write_tracking_id(workbook, worksheet, workbook_path, row, tracking_column, tracking_id):
    """Persist one row's TrackingId; SMTP must not run unless this returns successfully."""
    worksheet.cell(row, tracking_column, tracking_id)
    try:
        workbook.save(workbook_path)
    except Exception as exc:
        raise RuntimeError(f"Failed to write TrackingId to mail_list.xlsx: {exc}") from exc


def build_click_tracked_html(body, tracking_id):
    """Convert plain-text URLs to tracked HTML anchors while preserving visible text."""
    text = str(body or "")
    pattern = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)
    output = []
    position = 0
    trailing_punctuation = ".,;:!?)]}"
    for match in pattern.finditer(text):
        output.append(html.escape(text[position:match.start()]))
        candidate = match.group(0)
        url = candidate.rstrip(trailing_punctuation)
        trailing = candidate[len(url):]
        if url:
            destination = f"{TRACKING_BASE_URL}/email/click/{tracking_id}?url={quote(url, safe='')}"
            output.append(f'<a href="{html.escape(destination, quote=True)}">{html.escape(url)}</a>')
        output.append(html.escape(trailing))
        position = match.end()
    output.append(html.escape(text[position:]))
    return "".join(output).replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>\n")


def create_backup(source):
    ensure_dirs()
    target = app_paths()["backup"] / f"mail_list_{datetime.now():%Y%m%d_%H%M%S_%f}.xlsx"
    shutil.copy2(source, target)
    return target


def pending_count():
    p = app_paths()["list"]
    if not p.exists():
        return 0
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    h = header_map(ws)
    if "status" not in h:
        wb.close()
        return 0
    count = sum(1 for row in range(2, ws.max_row + 1) if str(ws.cell(row, h["status"]).value or "").strip().lower() == "pending")
    wb.close()
    return count


def safe_cell(ws, row, column):
    if not column:
        return ""
    return ws.cell(row, column).value


def row_value(values, column):
    if not column:
        return ""
    index = column - 1
    return values[index] if 0 <= index < len(values) else ""


def display_name_for_row(ws, h, row):
    if h.get("name"):
        return str(safe_cell(ws, row, h["name"]) or "")
    first = str(safe_cell(ws, row, h.get("first_name")) or "")
    last = str(safe_cell(ws, row, h.get("last_name")) or "")
    return f"{first} {last}".strip()


def page_bounds(total, page, page_size=PAGE_SIZE):
    total_pages = max(1, (max(0, total) + page_size - 1) // page_size)
    page = min(max(1, int(page or 1)), total_pages)
    start = (page - 1) * page_size
    end = min(start + page_size, max(0, total))
    return page, total_pages, start, end


def sortable_text(value):
    return str(value or "").casefold()


def row_matches(row_values, query):
    query = str(query or "").strip().casefold()
    return not query or query in " ".join(str(value or "") for value in row_values).casefold()


def email_row_values(ws, h, row):
    return (
        display_name_for_row(ws, h, row),
        str(safe_cell(ws, row, h.get("email")) or ""),
        str(safe_cell(ws, row, h.get("status")) or "").strip(),
        str(safe_cell(ws, row, h.get("sentdate")) or ""),
        str(safe_cell(ws, row, h.get("result")) or ""),
    )


def email_row_values_from_values(values, h):
    if h.get("name"):
        display_name = str(row_value(values, h["name"]) or "")
    else:
        first = str(row_value(values, h.get("first_name")) or "")
        last = str(row_value(values, h.get("last_name")) or "")
        display_name = f"{first} {last}".strip()
    return (
        display_name,
        str(row_value(values, h.get("email")) or ""),
        str(row_value(values, h.get("status")) or "").strip(),
        str(row_value(values, h.get("sentdate")) or ""),
        str(row_value(values, h.get("result")) or ""),
    )


def email_counts_from_sheet(ws, h):
    counts = {"total": 0, "sent": 0, "pending": 0, "failed": 0}
    for values in ws.iter_rows(min_row=2, values_only=True):
        status = str(row_value(values, h.get("status")) or "").strip()
        counts["total"] += 1
        key = status.lower()
        if key in counts:
            counts[key] += 1
    return counts


def email_grid_page(excel_path=None, page=1, page_size=PAGE_SIZE, search="", sort_key="", sort_reverse=False):
    p = Path(excel_path) if excel_path else app_paths()["list"]
    if not p.exists():
        return {"rows": [], "counts": {"total": 0, "sent": 0, "pending": 0, "failed": 0}, "page": 1, "total_pages": 1, "filtered_total": 0}
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    h = header_map(ws)
    counts = {"total": 0, "sent": 0, "pending": 0, "failed": 0}
    sort_indexes = {"name": 0, "email": 1, "status": 2, "sentdate": 3, "result": 4}
    query = str(search or "").strip()
    sort_index = sort_indexes.get(sort_key)
    if query or sort_index is not None:
        filtered = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            status = str(row_value(values, h.get("status")) or "").strip()
            counts["total"] += 1
            key = status.lower()
            if key in counts:
                counts[key] += 1
            row = email_row_values_from_values(values, h)
            if row_matches(row, query):
                filtered.append(row)
        if sort_index is not None:
            filtered.sort(key=lambda item: sortable_text(item[sort_index]), reverse=bool(sort_reverse))
        current_page, total_pages, start, end = page_bounds(len(filtered), page, page_size)
        rows = filtered[start:end]
        filtered_total = len(filtered)
    else:
        requested_page = max(1, int(page or 1))
        start = (requested_page - 1) * page_size
        end = start + page_size
        rows = []
        for index, values in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            status = str(row_value(values, h.get("status")) or "").strip()
            counts["total"] += 1
            key = status.lower()
            if key in counts:
                counts[key] += 1
            if start <= index < end:
                rows.append(email_row_values_from_values(values, h))
        filtered_total = counts["total"]
        current_page, total_pages, start, end = page_bounds(filtered_total, page, page_size)
        if current_page != requested_page:
            rows = []
            for index, values in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if start <= index < end:
                    rows.append(email_row_values_from_values(values, h))
    wb.close()
    return {"rows": rows, "counts": counts, "page": current_page, "total_pages": total_pages, "filtered_total": filtered_total}


def email_grid_data(excel_path=None):
    page = email_grid_page(excel_path, page=1, page_size=PAGE_SIZE)
    return page["rows"], page["counts"]


def dashboard_excel_summary(page=1, page_size=PAGE_SIZE, search="", sort_key="sender", sort_reverse=False):
    counts = {"total": 0, "sent": 0, "pending": 0, "failed": 0}
    today_sent = today_failed = weekly_sent = monthly_sent = 0
    per = {}
    p = app_paths()["list"]
    if p.exists():
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        h = header_map(ws)
        now = datetime.now()
        for values in ws.iter_rows(min_row=2, values_only=True):
            status = str(row_value(values, h.get("status")) or "")
            key = status.lower()
            counts["total"] += 1
            if key in counts:
                counts[key] += 1
            sender_key = h.get("sender_email") or h.get("sender_mail")
            sender = str(row_value(values, sender_key) or "") if sender_key else ""
            entry = per.setdefault(sender, {"sent_today": 0, "pending": 0})
            if key == "pending":
                entry["pending"] += 1
            date = row_value(values, h.get("sentdate")) if h.get("sentdate") else None
            is_today = isinstance(date, datetime) and date.date() == now.date()
            is_week = isinstance(date, datetime) and 0 <= (now.date() - date.date()).days < 7
            is_month = isinstance(date, datetime) and date.year == now.year and date.month == now.month
            if key == "sent":
                entry["sent_today"] += int(is_today)
                today_sent += int(is_today)
                weekly_sent += int(is_week)
                monthly_sent += int(is_month)
            if key == "failed" and is_today:
                today_failed += 1
        wb.close()
    rows = [(sender, value["sent_today"], value["pending"]) for sender, value in per.items()]
    query = str(search or "").strip()
    if query:
        rows = [row for row in rows if row_matches(row, query)]
    sort_indexes = {"sender": 0, "sent": 1, "pending": 2}
    sort_index = sort_indexes.get(sort_key)
    if sort_index is not None:
        rows.sort(key=lambda item: sortable_text(item[sort_index]), reverse=bool(sort_reverse))
    current_page, total_pages, start, end = page_bounds(len(rows), page, page_size)
    return {
        "counts": counts,
        "today_sent": today_sent,
        "today_failed": today_failed,
        "weekly_sent": weekly_sent,
        "monthly_sent": monthly_sent,
        "per_rows": rows[start:end],
        "per_page": current_page,
        "per_total_pages": total_pages,
        "per_filtered_total": len(rows),
    }


def send_pending(limit, wait_between=True, progress=None, excel_path=None, attachment_ids=None, attachment_source="Attachment Library", local_attachment_paths=None):
    paths = app_paths()
    if excel_path:
        paths["list"] = Path(excel_path)
    if limit < 1:
        raise ValueError("Email count must be at least 1.")
    if not paths["list"].exists():
        raise FileNotFoundError(f"Excel file not found: {paths['list']}")
    use_local_attachments = str(attachment_source or "Attachment Library") == "Local File"
    selected_attachments = [] if use_local_attachments else resolve_selected_attachments(attachment_ids)
    local_attachment_paths = [str(value) for value in (local_attachment_paths or []) if str(value or "").strip()]
    local_attachments = []
    if use_local_attachments:
        for value in local_attachment_paths:
            file_path = Path(value)
            if file_path.exists() and file_path.is_file():
                local_attachments.append(file_path)
            else:
                local_attachment_warning("Local attachment file missing; skipped.", path=str(file_path))
    wb = load_workbook(paths["list"])
    ws = wb.active
    h = validate_sheet(ws)
    cfg = load_config()
    if cfg.get("backup_enabled", True):
        create_backup(paths["list"])
    wb.save(paths["list"])
    tracking_column = ensure_tracking_column(wb, ws, paths["list"])
    h["trackingid"] = tracking_column
    rows = [r for r in range(2, ws.max_row + 1) if str(ws.cell(r, h["status"]).value or "").strip().lower() == "pending"][:limit]
    sent = failed = 0
    headers = [cell.value for cell in ws[1]]
    default_sender_name = str(load_config().get("default_sender_name") or "The Power People").strip()
    accounts = account_service()
    smtp = None
    active_sender = None
    try:
      for pos, row in enumerate(rows):
        started = time.monotonic()
        email = str(ws.cell(row, h["email"]).value or "").strip()
        subject = ""
        error = ""
        sender_email = ""
        tracking_id = ""
        try:
            values = [ws.cell(row, column).value for column in range(1, ws.max_column + 1)]
            context = PlaceholderService.create_context(headers, values)
            subject = PlaceholderService.render(ws.cell(row, h["subject"]).value, context).strip()
            body = PlaceholderService.render(ws.cell(row, h["body"]).value, context).strip()
            project_name = str(
                context.get("project_name")
                or context.get("project name")
                or context.get("project")
                or ""
            ).strip()
            sender_name = str(ws.cell(row, h["sender_name"]).value or "").strip() if h.get("sender_name") else ""
            sender_name = PlaceholderService.render(sender_name, context).strip() or default_sender_name
            sender_column = h.get("sender_email") or h.get("sender_mail")
            sender_email = str(ws.cell(row, sender_column).value or "").strip().lower() if sender_column else ""
            smtp_configuration = accounts.smtp_configuration(sender_email)
            if not smtp_configuration:
                raise ValueError("Sender Account Not Configured")
            if not validate_recipient_email(email):
                raise ValueError("Invalid email address")
            if not subject:
                raise ValueError("Subject is empty")
            if not body:
                raise ValueError("Body is empty")
            tracking_id = str(uuid.uuid4())
            write_tracking_id(wb, ws, paths["list"], row, tracking_column, tracking_id)
            if not use_local_attachments:
                register_attachment_mapping(tracking_id, selected_attachments)
            generated_message_id = generate_message_id()
            msg = EmailMessage()
            msg["From"] = f"{sender_name} <{AccountService.sender_address(smtp_configuration)}>"
            msg["To"] = email
            msg["Subject"] = subject
            if "Message-ID" not in msg:
                msg["Message-ID"] = generated_message_id
            msg.set_content(body)
            pixel_url = f"{TRACKING_BASE_URL}/email/open/{tracking_id}"
            html_body = build_click_tracked_html(body, tracking_id)
            attachment_html, _download_urls = build_attachment_links_html(tracking_id, selected_attachments) if not use_local_attachments else ("", [])
            email_html = f"{attachment_html}<hr>{html_body}" if attachment_html else html_body
            msg.add_alternative(f'<html><body>{email_html}<img src="{pixel_url}" width="1" height="1" style="display:none;" alt=""></body></html>', subtype="html")
            if use_local_attachments:
                for file_path in local_attachments:
                    attach_local_file(msg, file_path)
            if active_sender != sender_email or smtp is None:
                if smtp is not None:
                    try: smtp.quit()
                    except Exception: pass
                smtp = accounts.connect_smtp(smtp_configuration, timeout=45)
                active_sender = sender_email
            attempts = max(1, int(cfg.get("retry_count", 1)) + 1)
            last_error = None
            for attempt in range(attempts):
                try:
                    smtp.send_message(msg)
                    print("REAL SEND FUNCTION REACHED", flush=True)
                    tracking_registration_debug_log(event="REAL SEND FUNCTION REACHED", recipient_mail=email, tracking_id=tracking_id, generated_message_id=generated_message_id)
                    tracking_registration_debug_log(event="SMTP send successful", tracking_id=tracking_id, generated_message_id=generated_message_id, recipient_mail=email)
                    tracking_registration_logger().info("SMTP send successful tracking_id=%s generated_message_id=%s", tracking_id, generated_message_id)
                    sent_time = datetime.now(timezone.utc).isoformat()
                    register_sent_email(
                        tracking_id=tracking_id,
                        message_id=generated_message_id,
                        sender_mail=sender_email,
                        recipient_mail=email,
                        mail_subject=subject,
                        project_name=project_name,
                        excel_file_path=str(paths["list"].resolve()),
                        sent_time=sent_time,
                    )
                    last_error = None
                    break
                except Exception as exc:
                    last_error = exc
                    if attempt + 1 < attempts:
                        time.sleep(2)
            if last_error:
                raise last_error
            ws.cell(row, h["status"], "Sent")
            ws.cell(row, h["sentdate"], datetime.now())
            ws.cell(row, h["result"], "Success")
            sent += 1
            status = "Sent"
        except Exception as exc:
            error = str(exc)
            ws.cell(row, h["status"], "Failed")
            ws.cell(row, h["sentdate"], datetime.now())
            ws.cell(row, h["result"], error[:1000])
            failed += 1
            status = "Failed"
        wb.save(paths["list"])
        if progress:
            progress(pos + 1, len(rows), status, email)
        if wait_between and pos < len(rows) - 1:
            minimum = max(0, int(cfg.get("random_delay_min", 5)))
            maximum = max(minimum, int(cfg.get("random_delay_max", 15)))
            time.sleep(random.randint(minimum, maximum))
    finally:
      if smtp is not None:
        try: smtp.quit()
        except Exception: pass
    wb.close()
    return {"requested": limit, "sent": sent, "failed": failed, "remaining": pending_count()}


def hidden_startupinfo():
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    si.wShowWindow = subprocess.SW_HIDE
    return si


def run_cmd(args):
    return subprocess.run(args, capture_output=True, text=True, startupinfo=hidden_startupinfo(), creationflags=subprocess.CREATE_NO_WINDOW)


def task_query():
    result = run_cmd(["schtasks", "/Query", "/TN", TASK_NAME, "/FO", "LIST", "/V"])
    if result.returncode != 0:
        return {"exists": False, "enabled": False, "next": "Not scheduled"}
    text = result.stdout
    enabled = False
    next_run = "Unknown"
    for line in text.splitlines():
        if line.lower().startswith("scheduled task state"):
            enabled = line.split(":", 1)[1].strip().lower() == "enabled"
        if line.lower().startswith("next run time"):
            next_run = line.split(":", 1)[1].strip()
    return {"exists": True, "enabled": enabled, "next": next_run}


def scheduled_exe():
    return install_dir() / "SendPendingEmails.exe"


def update_task(cfg, enable=True):
    exe = scheduled_exe()
    if not exe.exists():
        raise FileNotFoundError(f"Installed sender not found: {exe}")
    command = f'"{exe}" --scheduled'
    result = run_cmd(["schtasks", "/Create", "/TN", TASK_NAME, "/TR", command, "/SC", "DAILY", "/ST", cfg["schedule_time"], "/F", "/RL", "LIMITED"])
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout).strip())
    ps = (
        "$s=New-Object -ComObject 'Schedule.Service';$s.Connect();"
        f"$t=$s.GetFolder('\\').GetTask('{TASK_NAME.replace("'", "''")}');"
        "$d=$t.Definition;$d.Settings.StartWhenAvailable=$true;"
        "$s.GetFolder('\\').RegisterTaskDefinition($t.Name,$d,6,$null,$null,3,$null)|Out-Null"
    )
    run_cmd(["powershell", "-NoProfile", "-WindowStyle", "Hidden", "-Command", ps])
    set_task_enabled(enable)


def set_task_enabled(enabled):
    action = "/Enable" if enabled else "/Disable"
    result = run_cmd(["schtasks", "/Change", "/TN", TASK_NAME, action])
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout).strip())


def delete_task():
    result = run_cmd(["schtasks", "/Delete", "/TN", TASK_NAME, "/F"])
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout).strip())


def create_shortcut(name, target):
    desktop = Path(os.environ.get("USERPROFILE", str(Path.home()))) / "Desktop"
    desktop.mkdir(parents=True, exist_ok=True)
    shortcut = desktop / f"{name}.lnk"
    ps = (
        "$w=New-Object -ComObject WScript.Shell;"
        f"$s=$w.CreateShortcut('{str(shortcut).replace("'", "''")}');"
        f"$s.TargetPath='{str(target).replace("'", "''")}';"
        f"$s.WorkingDirectory='{str(target.parent).replace("'", "''")}';"
        "$s.Save()"
    )
    result = run_cmd(["powershell", "-NoProfile", "-WindowStyle", "Hidden", "-Command", ps])
    if result.returncode != 0 or not shortcut.exists():
        raise RuntimeError(f"Could not create shortcut: {shortcut}")
    return shortcut


def setup_app():
    root = tk.Tk(); root.withdraw()
    try:
        dest = install_dir(); dest.mkdir(parents=True, exist_ok=True)
        for folder in ("config", "backup"):
            (dest / folder).mkdir(exist_ok=True)
        source = exe_dir()
        for name in ("SendPendingEmails.exe", "ConfigureSchedule.exe", "VerifyInstallation.exe"):
            src = source / name
            if not src.exists():
                raise FileNotFoundError(f"Deployment file is missing: {name}")
            shutil.copy2(src, dest / name)
        env_source = source / ".env"
        if not env_source.exists():
            env_source = source / ".env.template"
        if not env_source.exists():
            raise FileNotFoundError(".env or .env.template is missing from deployment package")
        shutil.copy2(env_source, dest / ".env")
        cfg = load_config(); save_config(cfg)
        create_shortcut("Send Pending Emails Now", dest / "SendPendingEmails.exe")
        create_shortcut("Configure Daily Email Schedule", dest / "ConfigureSchedule.exe")
        create_shortcut("Verify Installation", dest / "VerifyInstallation.exe")
        update_task(cfg, True)
        messagebox.showinfo(APP_NAME, f"Installation completed successfully.\n\nInstalled to:\n{dest}\n\nThree Desktop shortcuts were created.")
    except Exception as exc:
        messagebox.showerror(APP_NAME, f"Installation failed:\n\n{exc}")
    finally:
        root.destroy()


class AttachmentMultiSelect:
    SOURCES = ("Attachment Library", "Local File")

    def __init__(self, parent, selected_ids=None, source="Attachment Library", local_paths=None):
        self.frame = ttk.LabelFrame(parent, text="Attachments (Optional)", padding=8)
        self.frame.pack(fill="x", pady=(8, 0))
        self.saved_ids = [str(value) for value in (selected_ids or [])]
        self.local_paths = [str(value) for value in (local_paths or []) if str(value or "").strip()]
        self.attachments = []
        self.variables = {}
        self.source = tk.StringVar(value=source if source in self.SOURCES else "Attachment Library")
        source_row = ttk.Frame(self.frame); source_row.pack(fill="x")
        ttk.Label(source_row, text="Attachment Source:").pack(side="left")
        self.source_box = ttk.Combobox(source_row, textvariable=self.source, values=self.SOURCES, state="readonly", width=22)
        self.source_box.pack(side="left", padx=8)
        self.source_box.bind("<<ComboboxSelected>>", lambda _event: self.update_source())
        self.library_frame = ttk.Frame(self.frame)
        self.local_frame = ttk.Frame(self.frame)
        self.button_text = tk.StringVar(value="Select Attachment(s)")
        self.button = ttk.Menubutton(self.library_frame, textvariable=self.button_text, width=30)
        self.menu = tk.Menu(self.button, tearoff=False)
        self.button.configure(menu=self.menu); self.button.pack(side="left", pady=(8, 0))
        ttk.Button(self.library_frame, text="Refresh", command=self.refresh).pack(side="left", padx=8, pady=(8, 0))
        self.status = tk.StringVar(value=""); ttk.Label(self.library_frame, textvariable=self.status).pack(side="left", padx=8, pady=(8, 0))
        local_buttons = ttk.Frame(self.local_frame); local_buttons.pack(fill="x", pady=(8, 4))
        ttk.Button(local_buttons, text="Browse Local File(s)", command=self.browse_local_files).pack(side="left")
        ttk.Button(local_buttons, text="Remove Selected", command=self.remove_selected_local_files).pack(side="left", padx=8)
        self.local_status = tk.StringVar(value="")
        ttk.Label(local_buttons, textvariable=self.local_status).pack(side="left", padx=8)
        self.local_tree = ttk.Treeview(self.local_frame, columns=("name", "size", "path"), show="headings", height=4)
        for key, title, width in (("name", "File Name", 180), ("size", "File Size", 80), ("path", "Full Local Path", 430)):
            self.local_tree.heading(key, text=title); self.local_tree.column(key, width=width, minwidth=60)
        self.local_tree.pack(fill="x")
        self.refresh_local_files()
        self.update_source()
        self.refresh()

    def update_source(self):
        if self.source.get() == "Local File":
            self.library_frame.pack_forget()
            self.local_frame.pack(fill="x")
        else:
            self.local_frame.pack_forget()
            self.library_frame.pack(fill="x")

    def refresh(self):
        self.status.set("Loading attachments…"); self.saved_ids = self.selected_ids()
        def worker():
            try:
                attachments = attachment_library_service().list_attachments()
                self.frame.after(0, self.loaded, attachments)
            except Exception as exc: self.frame.after(0, self.failed, str(exc))
        threading.Thread(target=worker, daemon=True).start()

    def loaded(self, attachments):
        self.attachments = attachments; self.variables = {}; self.menu.delete(0, "end")
        for attachment in attachments:
            variable = tk.BooleanVar(value=attachment["id"] in self.saved_ids); self.variables[attachment["id"]] = variable
            self.menu.add_checkbutton(label=attachment["file_name"], variable=variable, command=self.update_text)
        self.status.set(f"{len(attachments)} available"); self.update_text()

    def failed(self, error): self.status.set(f"Attachment server unavailable: {error}")

    def selected_ids(self):
        if not self.variables: return list(self.saved_ids)
        return [attachment_id for attachment_id, variable in self.variables.items() if variable.get()]

    def selected_source(self):
        return self.source.get()

    def selected_local_paths(self):
        return list(self.local_paths)

    def update_text(self):
        count = len(self.selected_ids()); self.button_text.set("Select Attachment(s)" if count == 0 else f"{count} attachment(s) selected")

    def browse_local_files(self):
        values = filedialog.askopenfilenames(parent=self.frame.winfo_toplevel(), title="Select local attachment file(s)")
        for value in values:
            if value and value not in self.local_paths:
                self.local_paths.append(value)
        self.refresh_local_files()

    def remove_selected_local_files(self):
        selected_paths = {self.local_tree.item(item, "values")[2] for item in self.local_tree.selection()}
        self.local_paths = [value for value in self.local_paths if value not in selected_paths]
        self.refresh_local_files()

    def refresh_local_files(self):
        for item in self.local_tree.get_children():
            self.local_tree.delete(item)
        missing = 0
        for value in self.local_paths:
            file_path = Path(value)
            if file_path.exists() and file_path.is_file():
                info = local_attachment_info(file_path)
                self.local_tree.insert("", "end", values=(info["name"], format_file_size(info["size"]), info["path"]))
            else:
                missing += 1
                self.local_tree.insert("", "end", values=(file_path.name, "Missing", str(file_path)))
        total = len(self.local_paths)
        self.local_status.set(f"{total} selected" if not missing else f"{total} selected, {missing} missing")


class SenderWindow:
    def __init__(self, parent=None):
        self.root = tk.Toplevel(parent) if parent else tk.Tk()
        self.root.title("Send Pending Emails")
        self.root.geometry("920x590")
        self.root.minsize(760, 480)
        self.sending = False
        self.batch_sent = 0
        self.batch_failed = 0
        self.page = 1
        self.sort_key = ""
        self.sort_reverse = False
        self.total_pages = 1
        self.selected_rows = set()
        self.restoring_selection = False
        outer = ttk.Frame(self.root, padding=16)
        outer.pack(fill="both", expand=True)
        ttk.Label(outer, text="Email Sending Dashboard", font=("Segoe UI", 17, "bold")).pack(anchor="w")
        picker = ttk.Frame(outer); picker.pack(fill="x", pady=(10, 0))
        ttk.Label(picker, text="Excel File:").pack(side="left")
        self.excel_file = tk.StringVar(value=str(app_paths()["list"]))
        ttk.Entry(picker, textvariable=self.excel_file).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(picker, text="Browse", command=self.browse).pack(side="left")
        self.attachment_selector = AttachmentMultiSelect(outer)
        search_bar = ttk.Frame(outer); search_bar.pack(fill="x", pady=(10, 0))
        ttk.Label(search_bar, text="Search:").pack(side="left")
        self.search = tk.StringVar()
        ttk.Entry(search_bar, textvariable=self.search, width=34).pack(side="left", padx=7)
        ttk.Button(search_bar, text="Search", command=lambda: self.refresh(reset_page=True)).pack(side="left")
        ttk.Button(search_bar, text="Clear", command=self.clear_search).pack(side="left", padx=7)
        counts = ttk.Frame(outer)
        counts.pack(fill="x", pady=(14, 12))
        self.count_vars = {key: tk.StringVar(value=f"{label}: 0") for key, label in (("total", "Total Emails"), ("sent", "Sent Emails"), ("pending", "Pending Emails"), ("failed", "Failed Emails"))}
        for key in ("total", "sent", "pending", "failed"):
            ttk.Label(counts, textvariable=self.count_vars[key], font=("Segoe UI", 10, "bold"), padding=(0, 0, 28, 0)).pack(side="left")
        columns = ("name", "email", "status", "sentdate", "result")
        self.grid = ttk.Treeview(outer, columns=columns, show="headings", height=17)
        for key, title, width in (("name", "Name", 150), ("email", "Email", 220), ("status", "Status", 90), ("sentdate", "Sent Date", 150), ("result", "Result", 240)):
            self.grid.heading(key, text=title, command=lambda value=key: self.sort_by(value))
            self.grid.column(key, width=width, minwidth=70)
        scroll = ttk.Scrollbar(outer, orient="vertical", command=self.grid.yview)
        self.grid.configure(yscrollcommand=scroll.set)
        self.grid.bind("<<TreeviewSelect>>", self.remember_selection)
        self.grid.pack(side="left", fill="both", expand=True, pady=(0, 108))
        scroll.pack(side="left", fill="y", pady=(0, 108))
        bottom = ttk.Frame(self.root, padding=(16, 6, 16, 12))
        bottom.place(relx=0, rely=1, relwidth=1, anchor="sw")
        paging = ttk.Frame(bottom)
        paging.pack(fill="x", pady=(0, 6))
        self.first_button = ttk.Button(paging, text="First", command=lambda: self.goto_page(1))
        self.first_button.pack(side="left")
        self.previous_button = ttk.Button(paging, text="Previous", command=lambda: self.goto_page(self.page - 1))
        self.previous_button.pack(side="left", padx=(6, 0))
        self.page_label = tk.StringVar(value="Page 1 of 1")
        ttk.Label(paging, textvariable=self.page_label).pack(side="left", padx=12)
        self.next_button = ttk.Button(paging, text="Next", command=lambda: self.goto_page(self.page + 1))
        self.next_button.pack(side="left")
        self.last_button = ttk.Button(paging, text="Last", command=lambda: self.goto_page(self.total_pages))
        self.last_button.pack(side="left", padx=(6, 0))
        self.filtered_label = tk.StringVar(value="0 rows")
        ttk.Label(paging, textvariable=self.filtered_label).pack(side="left", padx=12)
        controls = ttk.Frame(bottom)
        controls.pack(fill="x")
        self.send_button = ttk.Button(controls, text="Send Pending Emails", command=self.start_send)
        self.send_button.pack(side="left")
        ttk.Button(controls, text="Refresh", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(controls, text="Close", command=self.root.destroy).pack(side="right")
        self.progress = ttk.Progressbar(controls, mode="determinate", length=230)
        self.progress.pack(side="right", padx=14)
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(controls, textvariable=self.status_var).pack(side="left", padx=14)
        self.refresh()

    def toast(self, message, error=False, duration=4500):
        toast = tk.Toplevel(self.root)
        toast.overrideredirect(True)
        toast.attributes("-topmost", True)
        color = "#a61b1b" if error else "#176b36"
        frame = tk.Frame(toast, bg=color, padx=18, pady=13)
        frame.pack(fill="both", expand=True)
        tk.Label(frame, text=message, bg=color, fg="white", font=("Segoe UI", 10, "bold"), wraplength=420, justify="left").pack()
        toast.update_idletasks()
        x = self.root.winfo_screenwidth() - toast.winfo_reqwidth() - 28
        y = self.root.winfo_screenheight() - toast.winfo_reqheight() - 70
        toast.geometry(f"+{x}+{y}")
        toast.after(duration, toast.destroy)

    def clear_search(self):
        self.search.set("")
        self.refresh(reset_page=True)

    def sort_by(self, key):
        if self.sort_key == key:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_key = key
            self.sort_reverse = False
        self.refresh(reset_page=True)

    @staticmethod
    def selection_key(values):
        return tuple(str(value) for value in values)

    def remember_selection(self, _event=None):
        if self.restoring_selection:
            return
        visible = {self.selection_key(self.grid.item(item, "values")) for item in self.grid.get_children()}
        selected = {self.selection_key(self.grid.item(item, "values")) for item in self.grid.selection()}
        self.selected_rows.difference_update(visible - selected)
        self.selected_rows.update(selected)

    def goto_page(self, page):
        self.remember_selection()
        try:
            self.page = int(page)
        except Exception:
            self.page = 1
        self.refresh()

    def update_paging_controls(self, filtered_total):
        self.page_label.set(f"Page {self.page} of {self.total_pages}")
        self.filtered_label.set(f"{filtered_total} rows")
        first_state = "disabled" if self.page <= 1 else "normal"
        last_state = "disabled" if self.page >= self.total_pages else "normal"
        self.first_button.configure(state=first_state)
        self.previous_button.configure(state=first_state)
        self.next_button.configure(state=last_state)
        self.last_button.configure(state=last_state)

    def refresh(self, reset_page=False):
        try:
            self.remember_selection()
            if reset_page:
                self.page = 1
            page = email_grid_page(self.excel_file.get(), page=self.page, page_size=PAGE_SIZE, search=self.search.get(), sort_key=self.sort_key, sort_reverse=self.sort_reverse)
            rows, counts = page["rows"], page["counts"]
            self.page = page["page"]
            self.total_pages = page["total_pages"]
            self.update_paging_controls(page["filtered_total"])
            for item in self.grid.get_children():
                self.grid.delete(item)
            selected_items = []
            for row in rows:
                item = self.grid.insert("", "end", values=row)
                if self.selection_key(row) in self.selected_rows:
                    selected_items.append(item)
            self.restoring_selection = True
            try:
                self.grid.selection_set(*selected_items)
            finally:
                self.restoring_selection = False
            labels = {"total": "Total Emails", "sent": "Sent Emails", "pending": "Pending Emails", "failed": "Failed Emails"}
            for key, label in labels.items():
                self.count_vars[key].set(f"{label}: {counts[key]}")
            return counts
        except Exception as exc:
            self.status_var.set(f"Refresh failed: {exc}")
            return {"total": 0, "sent": 0, "pending": 0, "failed": 0}

    def start_send(self):
        if self.sending:
            return
        counts = self.refresh()
        if counts["pending"] == 0:
            self.toast("There are no pending emails.")
            return
        value = simpledialog.askinteger(APP_NAME, "How many emails do you want to send now?", parent=self.root, minvalue=1)
        if value is None:
            return
        self.sending = True
        self.batch_sent = self.batch_failed = 0
        self.send_button.configure(state="disabled")
        self.progress.configure(maximum=min(value, counts["pending"]), value=0)
        self.status_var.set("Preparing to send…")
        excel_file = self.excel_file.get()
        attachment_source = self.attachment_selector.selected_source()
        attachment_ids = self.attachment_selector.selected_ids()
        local_attachment_paths = self.attachment_selector.selected_local_paths()
        if attachment_source == "Local File":
            missing = [path for path in local_attachment_paths if not Path(path).exists()]
            if missing:
                self.sending = False
                self.send_button.configure(state="normal")
                self.status_var.set("Ready")
                messagebox.showwarning(APP_NAME, "One or more selected local attachment files are missing:\n\n" + "\n".join(missing[:10]), parent=self.root)
                return

        def progress(number, total, result, email):
            if result == "Sent":
                self.batch_sent += 1
            else:
                self.batch_failed += 1
            self.root.after(0, self.on_progress, number, total, result, email)

        def worker():
            try:
                result = send_pending(value, True, progress, excel_file, attachment_ids, attachment_source, local_attachment_paths)
                self.root.after(0, self.on_complete, result)
            except PermissionError:
                message = "mail_list.xlsx is open or locked. Close the Excel file, then try again."
                self.root.after(0, self.on_error, message)
            except Exception as exc:
                self.root.after(0, self.on_error, str(exc))
        threading.Thread(target=worker, daemon=True).start()

    def on_progress(self, number, total, result, email):
        self.progress.configure(value=number, maximum=max(total, 1))
        counts = self.refresh()
        if result == "Sent":
            if counts["pending"] == 0:
                text = f"All {self.batch_sent} emails have been sent successfully."
            else:
                sent_word = "email" if self.batch_sent == 1 else "emails"
                pending_word = "email is" if counts["pending"] == 1 else "emails are"
                text = f"{self.batch_sent} {sent_word} sent successfully. {counts['pending']} {pending_word} pending."
            self.toast(text)
        else:
            self.toast(f"Failed to send email to {email}.", error=True)
        self.status_var.set(f"Processed {number} of {total}: {email}")

    def on_complete(self, result):
        self.sending = False
        self.send_button.configure(state="normal")
        counts = self.refresh()
        if result["failed"]:
            self.status_var.set(f"Finished: {result['sent']} sent, {result['failed']} failed")
        else:
            self.status_var.set(f"Finished: {result['sent']} sent successfully")
        if result["sent"] == 0 and result["failed"] == 0:
            self.toast("No pending emails were selected.")

    def on_error(self, message):
        self.sending = False
        self.send_button.configure(state="normal")
        self.refresh()
        self.status_var.set("Sending stopped")
        self.toast(message, error=True, duration=7000)

    def run(self):
        self.root.mainloop()

    def browse(self):
        value = filedialog.askopenfilename(parent=self.root, title="Select mail_list.xlsx", filetypes=[("Excel files", "*.xlsx")])
        if value: self.excel_file.set(value); self.refresh(reset_page=True)


def manual_send():
    SenderWindow().run()


class ScheduleWindow:
    def __init__(self):
        self.root = tk.Tk(); self.root.title("Configure Daily Email Schedule"); self.root.geometry("610x470"); self.root.resizable(False, False)
        self.cfg = load_config()
        frame = ttk.Frame(self.root, padding=20); frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Daily Email Schedule", font=("Segoe UI", 17, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 18))
        self.info = tk.StringVar(); ttk.Label(frame, textvariable=self.info, justify="left", font=("Segoe UI", 10)).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 20))
        ttk.Separator(frame).grid(row=2, column=0, columnspan=3, sticky="ew", pady=8)
        ttk.Label(frame, text="Emails per day:").grid(row=3, column=0, sticky="w", pady=8)
        self.limit = tk.StringVar(value=str(self.cfg["daily_limit"])); ttk.Entry(frame, textvariable=self.limit, width=16).grid(row=3, column=1, sticky="w")
        ttk.Label(frame, text="Daily time (HH:MM):").grid(row=4, column=0, sticky="w", pady=8)
        self.when = tk.StringVar(value=self.cfg["schedule_time"]); ttk.Entry(frame, textvariable=self.when, width=16).grid(row=4, column=1, sticky="w")
        ttk.Label(frame, text="Default sender name:").grid(row=5, column=0, sticky="w", pady=8)
        self.sender_name = tk.StringVar(value=self.cfg.get("default_sender_name", "The Power People")); ttk.Entry(frame, textvariable=self.sender_name, width=30).grid(row=5, column=1, sticky="w")
        buttons = ttk.Frame(frame); buttons.grid(row=6, column=0, columnspan=3, sticky="w", pady=22)
        ttk.Button(buttons, text="Save Changes", command=self.save).pack(side="left", padx=(0, 7))
        ttk.Button(buttons, text="Enable Schedule", command=lambda: self.enable(True)).pack(side="left", padx=7)
        ttk.Button(buttons, text="Disable Schedule", command=lambda: self.enable(False)).pack(side="left", padx=7)
        ttk.Button(buttons, text="Delete Schedule", command=self.delete).pack(side="left", padx=7)
        ttk.Button(frame, text="Close", command=self.root.destroy).grid(row=7, column=0, sticky="w")
        ttk.Label(frame, text="Note: missed runs start when Windows next allows the task. Running while logged off may require Windows credentials and administrator policy.", wraplength=555, foreground="#555").grid(row=8, column=0, columnspan=3, sticky="w", pady=(25, 0))
        self.refresh()
    def refresh(self):
        state = task_query()
        status = "Enabled" if state["enabled"] else ("Disabled" if state["exists"] else "Not Created")
        self.info.set(f"Current Daily Email Limit:  {self.cfg['daily_limit']}\nCurrent Scheduled Time:  {self.cfg['schedule_time']}\nScheduled Task Status:  {status}\nNext Scheduled Run Time:  {state['next']}")
    def values(self):
        limit = int(self.limit.get())
        if limit < 1: raise ValueError("Emails per day must be at least 1.")
        datetime.strptime(self.when.get().strip(), "%H:%M")
        sender_name = self.sender_name.get().strip()
        if not sender_name: raise ValueError("Default sender name cannot be empty.")
        return limit, self.when.get().strip(), sender_name
    def save(self):
        try:
            limit, when, sender_name = self.values(); self.cfg.update({"daily_limit": limit, "schedule_time": when, "default_sender_name": sender_name}); save_config(self.cfg); update_task(self.cfg, True); self.refresh(); messagebox.showinfo(APP_NAME, "Configuration saved and schedule enabled.")
        except Exception as exc: messagebox.showerror(APP_NAME, str(exc))
    def enable(self, enabled):
        try:
            if enabled and not task_query()["exists"]: update_task(self.cfg, True)
            else: set_task_enabled(enabled)
            self.refresh()
        except Exception as exc: messagebox.showerror(APP_NAME, str(exc))
    def delete(self):
        if not messagebox.askyesno(APP_NAME, "Delete the daily scheduled task?"): return
        try: delete_task(); self.refresh()
        except Exception as exc: messagebox.showerror(APP_NAME, str(exc))
    def run(self): self.root.mainloop()


class MailAccountEditor:
    def __init__(self, parent, account=None):
        self.account = account or {}
        self.result = None
        self.window = tk.Toplevel(parent); self.window.title("Edit Mail Account" if account else "Add Mail Account"); self.window.geometry("620x720"); self.window.resizable(False, False); self.window.transient(parent); self.window.grab_set()
        frame = ttk.Frame(self.window, padding=20); frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Mail Provider Account", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 16))
        self.provider = tk.StringVar(value=self.account.get("provider", "Gmail"))
        self.display_name = tk.StringVar(value=self.account.get("display_name", self.account.get("name", "")))
        self.email = tk.StringVar(value=self.account.get("email", ""))
        self.sender_alias = tk.StringVar(value=self.account.get("sender_alias", ""))
        self.host = tk.StringVar(value=self.account.get("smtp_host", "smtp.gmail.com"))
        self.port = tk.StringVar(value=str(self.account.get("smtp_port", 587)))
        self.encryption = tk.StringVar(value=self.account.get("encryption", "STARTTLS"))
        self.imap_host = tk.StringVar(value=self.account.get("imap_host", "imap.gmail.com"))
        self.imap_port = tk.StringVar(value=str(self.account.get("imap_port", 993)))
        self.imap_encryption = tk.StringVar(value=self.account.get("imap_encryption", "SSL/TLS"))
        self.password = tk.StringVar()
        self.enabled = tk.BooleanVar(value=self.account.get("enabled", True))
        labels = (("Mail Provider:", 1), ("Display Name:", 2), ("Email Address:", 3), ("Sender Alias (Optional):", 4), ("SMTP Host:", 6), ("SMTP Port:", 7), ("SMTP Encryption:", 8), ("IMAP Host:", 9), ("IMAP Port:", 10), ("IMAP Encryption:", 11), ("Password / App Password:", 12))
        for text, row in labels: ttk.Label(frame, text=text).grid(row=row, column=0, sticky="w", pady=7)
        self.provider_box = ttk.Combobox(frame, textvariable=self.provider, values=AccountService.PROVIDERS, state="readonly", width=32); self.provider_box.grid(row=1, column=1, sticky="w"); self.provider_box.bind("<<ComboboxSelected>>", self.provider_changed)
        ttk.Entry(frame, textvariable=self.display_name, width=38).grid(row=2, column=1, sticky="w")
        ttk.Entry(frame, textvariable=self.email, width=38).grid(row=3, column=1, sticky="w")
        self.sender_alias_entry = ttk.Entry(frame, textvariable=self.sender_alias, width=38); self.sender_alias_entry.grid(row=4, column=1, sticky="w")
        ttk.Label(frame, text="For Gmail verified 'Send mail as' aliases only.", foreground="#555").grid(row=5, column=1, sticky="w", pady=(0, 4))
        self.host_entry = ttk.Entry(frame, textvariable=self.host, width=38); self.host_entry.grid(row=6, column=1, sticky="w")
        self.port_entry = ttk.Entry(frame, textvariable=self.port, width=15); self.port_entry.grid(row=7, column=1, sticky="w")
        self.encryption_box = ttk.Combobox(frame, textvariable=self.encryption, values=AccountService.ENCRYPTION_TYPES, state="readonly", width=32); self.encryption_box.grid(row=8, column=1, sticky="w")
        self.imap_host_entry = ttk.Entry(frame, textvariable=self.imap_host, width=38); self.imap_host_entry.grid(row=9, column=1, sticky="w")
        self.imap_port_entry = ttk.Entry(frame, textvariable=self.imap_port, width=15); self.imap_port_entry.grid(row=10, column=1, sticky="w")
        self.imap_encryption_box = ttk.Combobox(frame, textvariable=self.imap_encryption, values=AccountService.ENCRYPTION_TYPES, state="readonly", width=32); self.imap_encryption_box.grid(row=11, column=1, sticky="w")
        ttk.Entry(frame, textvariable=self.password, show="*", width=38).grid(row=12, column=1, sticky="w")
        password_note = "Leave empty while editing to keep the stored password." if account else "Gmail requires a Google App Password."
        ttk.Label(frame, text=password_note, foreground="#555").grid(row=13, column=1, sticky="w", pady=(0, 8))
        ttk.Checkbutton(frame, text="Enabled", variable=self.enabled).grid(row=14, column=1, sticky="w", pady=8)
        buttons = ttk.Frame(frame); buttons.grid(row=15, column=0, columnspan=2, sticky="e", pady=(18, 0))
        ttk.Button(buttons, text="Save", command=self.save).pack(side="left", padx=5); ttk.Button(buttons, text="Cancel", command=self.window.destroy).pack(side="left", padx=5)
        self.apply_provider_state(reset_defaults=False)
        self.window.wait_window()
    def provider_changed(self, _event=None): self.apply_provider_state(reset_defaults=True)
    def apply_provider_state(self, reset_defaults=False):
        gmail = self.provider.get() == "Gmail"
        if gmail:
            self.host.set("smtp.gmail.com"); self.port.set("587"); self.encryption.set("STARTTLS"); self.imap_host.set("imap.gmail.com"); self.imap_port.set("993"); self.imap_encryption.set("SSL/TLS")
        elif reset_defaults:
            self.host.set(""); self.port.set("587"); self.encryption.set("STARTTLS"); self.imap_host.set(""); self.imap_port.set("993"); self.imap_encryption.set("SSL/TLS")
        self.host_entry.configure(state="disabled" if gmail else "normal")
        self.port_entry.configure(state="disabled" if gmail else "normal")
        self.encryption_box.configure(state="disabled" if gmail else "readonly")
        self.imap_host_entry.configure(state="disabled" if gmail else "normal")
        self.imap_port_entry.configure(state="disabled" if gmail else "normal")
        self.imap_encryption_box.configure(state="disabled" if gmail else "readonly")
        self.sender_alias_entry.configure(state="normal" if gmail else "disabled")
    def save(self):
        try:
            provider = self.provider.get()
            host = "smtp.gmail.com" if provider == "Gmail" else self.host.get().strip()
            port = 587 if provider == "Gmail" else self.port.get().strip()
            encryption = "STARTTLS" if provider == "Gmail" else self.encryption.get()
            imap_host = "imap.gmail.com" if provider == "Gmail" else self.imap_host.get().strip()
            imap_port = 993 if provider == "Gmail" else self.imap_port.get().strip()
            imap_encryption = "SSL/TLS" if provider == "Gmail" else self.imap_encryption.get()
            if not host: raise ValueError("SMTP Host cannot be empty.")
            if not str(port).isdigit(): raise ValueError("SMTP Port must be numeric.")
            if not imap_host: raise ValueError("IMAP Host cannot be empty.")
            if not str(imap_port).isdigit(): raise ValueError("IMAP Port must be numeric.")
            if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", self.email.get().strip()): raise ValueError("A valid email address is required.")
            self.result = {"name": self.display_name.get().strip(), "email": self.email.get().strip(), "sender_alias": self.sender_alias.get().strip() if provider == "Gmail" else "", "password": self.password.get(), "enabled": self.enabled.get(), "provider": provider, "smtp_host": host, "smtp_port": port, "encryption": encryption, "imap_host": imap_host, "imap_port": imap_port, "imap_encryption": imap_encryption}
            self.window.destroy()
        except ValueError as exc: messagebox.showerror(APP_NAME, str(exc), parent=self.window)


class AccountWindow:
    def __init__(self, parent=None):
        self.root = tk.Toplevel(parent) if parent else tk.Tk()
        self.root.title("Mail Account Setup")
        self.root.geometry("1050x460")
        self.service = account_service()
        frame = ttk.Frame(self.root, padding=16); frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Mail Account Management", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 12))
        self.tree = ttk.Treeview(frame, columns=("provider", "name", "email", "host", "port", "encryption", "status"), show="headings")
        for key, title, width in (("provider", "Provider", 110), ("name", "Display Name", 150), ("email", "Email Address", 210), ("host", "SMTP Host", 180), ("port", "Port", 60), ("encryption", "Encryption", 90), ("status", "Status", 80)):
            self.tree.heading(key, text=title); self.tree.column(key, width=width)
        self.tree.pack(fill="both", expand=True)
        bar = ttk.Frame(frame); bar.pack(fill="x", pady=(12, 0))
        for text, command in (("Add", self.add), ("Edit", self.edit), ("Delete", self.delete), ("Enable", lambda: self.toggle(True)), ("Disable", lambda: self.toggle(False)), ("Test SMTP Connection", self.test)):
            ttk.Button(bar, text=text, command=command).pack(side="left", padx=(0, 7))
        ttk.Button(bar, text="Close", command=self.root.destroy).pack(side="right")
        self.refresh()
    def selected(self):
        selected = self.tree.selection()
        return self.service.find(self.tree.item(selected[0], "values")[2]) if selected else None
    def refresh(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        for account in self.service.list_accounts():
            self.tree.insert("", "end", values=(account["provider"], account["display_name"], account["email"], account["smtp_host"], account["smtp_port"], account["encryption"], "Enabled" if account.get("enabled", True) else "Disabled"))
    def account_form(self, existing=None):
        values = MailAccountEditor(self.root, existing).result
        if values:
            self.service.save_account(original_email=existing["email"] if existing else None, **values); self.refresh()
    def add(self):
        try: self.account_form()
        except Exception as exc: messagebox.showerror(APP_NAME, str(exc), parent=self.root)
    def edit(self):
        value = self.selected()
        if not value: return
        try: self.account_form(value)
        except Exception as exc: messagebox.showerror(APP_NAME, str(exc), parent=self.root)
    def delete(self):
        value = self.selected()
        if value and messagebox.askyesno(APP_NAME, f"Delete {value['email']} permanently?", parent=self.root):
            try: self.service.delete_account(value["email"]); self.refresh()
            except Exception as exc: messagebox.showerror(APP_NAME, str(exc), parent=self.root)
    def toggle(self, enabled):
        value = self.selected()
        if value: self.service.set_enabled(value["email"], enabled); self.refresh()
    def test(self):
        value = self.selected()
        if not value: return
        try: self.service.test_smtp(value["email"]); messagebox.showinfo(APP_NAME, "SMTP connection successful.", parent=self.root)
        except Exception as exc: messagebox.showerror(APP_NAME, str(exc), parent=self.root)


def load_schedules():
    path = app_paths()["schedules"]
    if not path.exists(): return []
    try: return json.loads(path.read_text(encoding="utf-8"))
    except Exception: return []


def save_schedules(items):
    path = app_paths()["schedules"]; path.parent.mkdir(parents=True, exist_ok=True); path.write_text(json.dumps(items, indent=2), encoding="utf-8")


def schedule_task_name(item):
    return f"Email Automation - {item['id']}"


def sync_schedule_task(item):
    name = schedule_task_name(item)
    if not item.get("enabled", True):
        run_cmd(["schtasks", "/Delete", "/TN", name, "/F"]); return
    days = ",".join(day[:3].upper() for day in item["days"])
    if not days: raise ValueError("Select at least one day.")
    exe = Path(sys.executable).resolve() if getattr(sys, "frozen", False) else Path(__file__).resolve()
    action = f'"{exe}" --run-schedule {item["id"]}'
    result = run_cmd(["schtasks", "/Create", "/TN", name, "/TR", action, "/SC", "WEEKLY", "/D", days, "/ST", item["time"], "/F", "/RL", "LIMITED"])
    if result.returncode != 0: raise RuntimeError((result.stderr or result.stdout).strip())
    ps = "$s=New-Object -ComObject 'Schedule.Service';$s.Connect();$t=$s.GetFolder('\\').GetTask('" + name.replace("'", "''") + "');$d=$t.Definition;$d.Settings.StartWhenAvailable=$true;$s.GetFolder('\\').RegisterTaskDefinition($t.Name,$d,6,$null,$null,3,$null)|Out-Null"
    run_cmd(["powershell", "-NoProfile", "-WindowStyle", "Hidden", "-Command", ps])


class ScheduleEditor:
    DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    def __init__(self, parent, current=None):
        self.current = current or {}
        self.result = None
        self.window = tk.Toplevel(parent)
        self.window.title("Edit Schedule" if current else "Add Schedule")
        self.window.geometry("820x760")
        self.window.resizable(False, False)
        self.window.transient(parent)
        self.window.grab_set()
        frame = ttk.Frame(self.window, padding=20)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Schedule Details", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 16))
        self.name = tk.StringVar(value=self.current.get("name", ""))
        self.excel = tk.StringVar(value=self.current.get("excel_file", ""))
        self.when = tk.StringVar(value=self.current.get("time", "09:00"))
        self.maximum = tk.StringVar(value=str(self.current.get("max_emails", 10)))
        self.enabled = tk.BooleanVar(value=self.current.get("enabled", True))
        ttk.Label(frame, text="Schedule Name:").grid(row=1, column=0, sticky="w", pady=7)
        ttk.Entry(frame, textvariable=self.name, width=48).grid(row=1, column=1, columnspan=2, sticky="ew", pady=7)
        ttk.Label(frame, text="Excel File:").grid(row=2, column=0, sticky="w", pady=7)
        ttk.Entry(frame, textvariable=self.excel, width=40).grid(row=2, column=1, sticky="ew", pady=7)
        ttk.Button(frame, text="Browse", command=self.browse).grid(row=2, column=2, padx=(8, 0), pady=7)
        ttk.Label(frame, text="Days of Week:").grid(row=3, column=0, sticky="nw", pady=(12, 7))
        day_frame = ttk.LabelFrame(frame, text="Select one or more days", padding=12)
        day_frame.grid(row=3, column=1, columnspan=2, sticky="ew", pady=(12, 7))
        saved_days = set(self.current.get("days", []))
        self.day_vars = {}
        for index, day in enumerate(self.DAYS):
            variable = tk.BooleanVar(value=day in saved_days)
            self.day_vars[day] = variable
            ttk.Checkbutton(day_frame, text=day, variable=variable).grid(row=index // 2, column=index % 2, sticky="w", padx=(0, 36), pady=4)
        attachment_host = ttk.Frame(frame); attachment_host.grid(row=4, column=0, columnspan=3, sticky="ew")
        self.attachment_selector = AttachmentMultiSelect(
            attachment_host,
            self.current.get("attachment_ids", []),
            self.current.get("attachment_source", "Attachment Library"),
            self.current.get("local_attachment_paths", []),
        )
        ttk.Label(frame, text="Time (HH:mm):").grid(row=5, column=0, sticky="w", pady=7)
        ttk.Entry(frame, textvariable=self.when, width=16).grid(row=5, column=1, sticky="w", pady=7)
        ttk.Label(frame, text="Maximum Emails:").grid(row=6, column=0, sticky="w", pady=7)
        ttk.Entry(frame, textvariable=self.maximum, width=16).grid(row=6, column=1, sticky="w", pady=7)
        ttk.Checkbutton(frame, text="Schedule Enabled", variable=self.enabled).grid(row=7, column=1, sticky="w", pady=10)
        buttons = ttk.Frame(frame)
        buttons.grid(row=8, column=0, columnspan=3, sticky="e", pady=(18, 0))
        ttk.Button(buttons, text="Save", command=self.save).pack(side="left", padx=5)
        ttk.Button(buttons, text="Cancel", command=self.window.destroy).pack(side="left", padx=5)
        frame.columnconfigure(1, weight=1)
        self.window.protocol("WM_DELETE_WINDOW", self.window.destroy)
        self.window.wait_window()

    def browse(self):
        value = filedialog.askopenfilename(parent=self.window, title="Select mail_list.xlsx", filetypes=[("Excel files", "*.xlsx")])
        if value:
            self.excel.set(value)

    def save(self):
        selected_days = [day for day in self.DAYS if self.day_vars[day].get()]
        if not selected_days:
            messagebox.showerror(APP_NAME, "Please select at least one day.", parent=self.window)
            return
        try:
            if not self.name.get().strip():
                raise ValueError("Schedule Name is required.")
            if not self.excel.get().strip():
                raise ValueError("Excel File is required.")
            datetime.strptime(self.when.get().strip(), "%H:%M")
            maximum = int(self.maximum.get())
            if maximum < 1:
                raise ValueError("Maximum Emails must be at least 1.")
            self.result = {
                "id": self.current.get("id", str(uuid.uuid4())),
                "name": self.name.get().strip(),
                "excel_file": self.excel.get().strip(),
                "days": selected_days,
                "time": self.when.get().strip(),
                "max_emails": maximum,
                "attachment_source": self.attachment_selector.selected_source(),
                "attachment_ids": self.attachment_selector.selected_ids(),
                "local_attachment_paths": self.attachment_selector.selected_local_paths(),
                "enabled": self.enabled.get(),
            }
            self.window.destroy()
        except ValueError as exc:
            messagebox.showerror(APP_NAME, str(exc), parent=self.window)


class MultiScheduleWindow:
    DAYS = ScheduleEditor.DAYS
    def __init__(self, parent=None):
        self.root = tk.Toplevel(parent) if parent else tk.Tk(); self.root.title("Daily Scheduling"); self.root.geometry("900x480")
        self.page = 1
        self.total_pages = 1
        self.page_size = tk.IntVar(value=PAGE_SIZE)
        self.sort_key = ""
        self.sort_reverse = False
        frame = ttk.Frame(self.root, padding=16); frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Recurring Email Schedules", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 12))
        self.tree = ttk.Treeview(frame, columns=("name", "file", "days", "time", "maximum", "status"), show="headings")
        for key, title, width in (("name", "Schedule Name", 140), ("file", "Excel File", 220), ("days", "Days", 190), ("time", "Time", 70), ("maximum", "Maximum", 70), ("status", "Status", 80)):
            self.tree.heading(key, text=title, command=lambda value=key:self.sort_by(value)); self.tree.column(key, width=width)
        self.tree.pack(fill="both", expand=True)
        paging=ttk.Frame(frame);paging.pack(fill="x",pady=(10,0))
        self.first_button=ttk.Button(paging,text="First",command=lambda:self.goto_page(1));self.first_button.pack(side="left")
        self.previous_button=ttk.Button(paging,text="Previous",command=lambda:self.goto_page(self.page-1));self.previous_button.pack(side="left",padx=(6,0))
        self.page_label=tk.StringVar(value="Page 1 of 1");ttk.Label(paging,textvariable=self.page_label).pack(side="left",padx=12)
        self.next_button=ttk.Button(paging,text="Next",command=lambda:self.goto_page(self.page+1));self.next_button.pack(side="left")
        self.last_button=ttk.Button(paging,text="Last",command=lambda:self.goto_page(self.total_pages));self.last_button.pack(side="left",padx=(6,18))
        ttk.Label(paging,text="Page Size:").pack(side="left")
        page_size_box=ttk.Combobox(paging,textvariable=self.page_size,values=(10,20,50,100),state="readonly",width=6)
        page_size_box.pack(side="left",padx=(6,0));page_size_box.bind("<<ComboboxSelected>>",lambda _event:self.change_page_size())
        bar=ttk.Frame(frame);bar.pack(fill="x",pady=12)
        for text,cmd in (("Add",self.add),("Edit",self.edit),("Delete",self.delete),("Enable",lambda:self.toggle(True)),("Disable",lambda:self.toggle(False)),("Duplicate",self.duplicate)):
            ttk.Button(bar,text=text,command=cmd).pack(side="left",padx=(0,7))
        ttk.Button(bar,text="Close",command=self.root.destroy).pack(side="right");self.refresh()
    def items(self): return load_schedules()
    def selected_id(self):
        values=self.tree.selection(); return self.tree.item(values[0],"tags")[0] if values else None
    def display_values(self,item):
        return (item["name"],item["excel_file"],", ".join(item["days"]),item["time"],item["max_emails"],"Enabled" if item.get("enabled",True) else "Disabled")
    def sorted_items(self):
        values=list(self.items())
        indexes={"name":0,"file":1,"days":2,"time":3,"maximum":4,"status":5}
        if self.sort_key in indexes:
            index=indexes[self.sort_key]
            values.sort(key=lambda item: sortable_text(self.display_values(item)[index]),reverse=self.sort_reverse)
        return values
    def sort_by(self,key):
        if self.sort_key==key:self.sort_reverse=not self.sort_reverse
        else:self.sort_key=key;self.sort_reverse=False
        self.page=1;self.refresh()
    def change_page_size(self):
        try:self.page_size.set(max(1,int(self.page_size.get())))
        except Exception:self.page_size.set(PAGE_SIZE)
        self.page=1;self.refresh()
    def goto_page(self,page):
        try:self.page=int(page)
        except Exception:self.page=1
        self.refresh()
    def update_paging_controls(self,total_items):
        size=max(1,int(self.page_size.get() or PAGE_SIZE))
        self.page,self.total_pages,start,end=page_bounds(total_items,self.page,size)
        self.page_label.set(f"Page {self.page} of {self.total_pages}")
        first_state="disabled" if self.page<=1 else "normal"
        last_state="disabled" if self.page>=self.total_pages else "normal"
        self.first_button.configure(state=first_state);self.previous_button.configure(state=first_state)
        self.next_button.configure(state=last_state);self.last_button.configure(state=last_state)
        return start,end
    def refresh(self):
        selected=self.selected_id()
        values=self.sorted_items()
        start,end=self.update_paging_controls(len(values))
        for row in self.tree.get_children(): self.tree.delete(row)
        for item in values[start:end]:
            row=self.tree.insert("","end",values=self.display_values(item),tags=(item["id"],))
            if selected and item["id"]==selected:self.tree.selection_set(row)
    def form(self, item=None):
        return ScheduleEditor(self.root, item).result
    def add(self):
        try:
            item=self.form()
            if item: values=self.items();values.append(item);save_schedules(values);sync_schedule_task(item);self.refresh()
        except Exception as exc:messagebox.showerror(APP_NAME,str(exc),parent=self.root)
    def edit(self):
        key=self.selected_id();values=self.items();old=next((x for x in values if x["id"]==key),None)
        if not old:return
        try:
            updated=self.form(old)
            if updated: values=[updated if x["id"]==key else x for x in values];save_schedules(values);sync_schedule_task(updated);self.refresh()
        except Exception as exc:messagebox.showerror(APP_NAME,str(exc),parent=self.root)
    def delete(self):
        key=self.selected_id()
        if not key or not messagebox.askyesno(APP_NAME,"Delete this schedule?",parent=self.root):return
        values=self.items();item=next(x for x in values if x["id"]==key);run_cmd(["schtasks","/Delete","/TN",schedule_task_name(item),"/F"]);save_schedules([x for x in values if x["id"]!=key]);self.refresh()
    def toggle(self,enabled):
        key=self.selected_id();values=self.items()
        for item in values:
            if item["id"]==key:item["enabled"]=enabled;sync_schedule_task(item)
        save_schedules(values);self.refresh()
    def duplicate(self):
        key=self.selected_id();values=self.items();old=next((x for x in values if x["id"]==key),None)
        if old:
            item=dict(old);item["id"]=str(uuid.uuid4());item["name"] += " Copy";values.append(item);save_schedules(values);sync_schedule_task(item);self.refresh()


class GlobalSettingsWindow:
    def __init__(self, parent):
        self.root=tk.Toplevel(parent);self.root.title("Settings");self.root.geometry("520x470");self.cfg=load_config();frame=ttk.Frame(self.root,padding=18);frame.pack(fill="both",expand=True)
        ttk.Label(frame,text="Global Settings",font=("Segoe UI",16,"bold")).grid(row=0,column=0,columnspan=2,sticky="w",pady=(0,15))
        fields=(("random_delay_min","Minimum Delay Seconds"),("random_delay_max","Maximum Delay Seconds"),("retry_count","Retry Count"),("default_sender_name","Default Sender Name"))
        self.vars={}
        for row,(key,label) in enumerate(fields,1):
            ttk.Label(frame,text=label+":").grid(row=row,column=0,sticky="w",pady=6);self.vars[key]=tk.StringVar(value=str(self.cfg.get(key,"")));ttk.Entry(frame,textvariable=self.vars[key],width=28).grid(row=row,column=1,sticky="w")
        self.backup=tk.BooleanVar(value=self.cfg.get("backup_enabled",True));ttk.Checkbutton(frame,text="Backup Enabled",variable=self.backup).grid(row=5,column=0,columnspan=2,sticky="w",pady=8)
        ttk.Label(frame,text="Default Theme:").grid(row=6,column=0,sticky="w");self.theme=tk.StringVar(value=self.cfg.get("theme","Light"));ttk.Combobox(frame,textvariable=self.theme,values=("Light","Dark"),state="readonly",width=25).grid(row=6,column=1,sticky="w")
        bar=ttk.Frame(frame);bar.grid(row=7,column=0,columnspan=2,sticky="w",pady=18)
        ttk.Button(bar,text="Save",command=self.save).pack(side="left",padx=(0,7))
        for text,key in (("Open Backup Folder","backup"),):
            ttk.Button(bar,text=text,command=lambda k=key:os.startfile(app_paths()[k])).pack(side="left",padx=(0,7))
        ttk.Button(frame,text="Check for Updates (Future)",command=lambda:messagebox.showinfo(APP_NAME,"Update service is future-ready.",parent=self.root)).grid(row=8,column=0,columnspan=2,sticky="w")
    def save(self):
        try:
            for key in ("random_delay_min","random_delay_max","retry_count"):self.cfg[key]=int(self.vars[key].get())
            self.cfg["default_sender_name"]=self.vars["default_sender_name"].get().strip();self.cfg["backup_enabled"]=self.backup.get();self.cfg["theme"]=self.theme.get();save_config(self.cfg);messagebox.showinfo(APP_NAME,"Settings saved.",parent=self.root)
        except Exception as exc:messagebox.showerror(APP_NAME,str(exc),parent=self.root)


class SettingsWindow:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent); self.root.title("Settings"); self.root.geometry("520x300"); self.root.resizable(False, False)
        frame = ttk.Frame(self.root, padding=24); frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Settings", font=("Segoe UI", 18, "bold")).pack(anchor="w", pady=(0, 22))
        ttk.Button(frame, text="Global Settings", command=lambda:GlobalSettingsWindow(self.root)).pack(fill="x", ipady=16, pady=(0, 14))
        ttk.Button(frame, text="Attachment Library", command=lambda:AttachmentLibraryWindow(self.root)).pack(fill="x", ipady=16)


class AttachmentLibraryWindow:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent); self.root.title("Attachment Library"); self.root.geometry("790x520")
        self.service = attachment_library_service(); self.attachments_by_item = {}
        frame = ttk.Frame(self.root, padding=18); frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Attachment Library", font=("Segoe UI", 17, "bold")).pack(anchor="w", pady=(0, 14))
        self.tree = ttk.Treeview(frame, columns=("name", "size", "date"), show="headings", height=16)
        for key, title, width in (("name", "File Name", 340), ("size", "File Size", 130), ("date", "Upload Date", 240)):
            self.tree.heading(key, text=title); self.tree.column(key, width=width)
        self.tree.pack(fill="both", expand=True)
        self.progress = ttk.Progressbar(frame, mode="determinate"); self.progress.pack(fill="x", pady=(12, 4))
        self.status = tk.StringVar(value="Ready"); ttk.Label(frame, textvariable=self.status).pack(anchor="w", pady=(0, 8))
        bar = ttk.Frame(frame); bar.pack(fill="x")
        self.upload_button = ttk.Button(bar, text="Upload", command=self.upload); self.upload_button.pack(side="left", padx=(0, 7))
        self.delete_button = ttk.Button(bar, text="Delete", command=self.delete); self.delete_button.pack(side="left", padx=7)
        self.refresh_button = ttk.Button(bar, text="Refresh", command=self.refresh); self.refresh_button.pack(side="left", padx=7)
        ttk.Button(bar, text="Close", command=self.root.destroy).pack(side="right")
        self.refresh()
    @staticmethod
    def display_size(size):
        value = float(size)
        for unit in ("B", "KB", "MB", "GB"):
            if value < 1024 or unit == "GB": return f"{value:.0f} {unit}" if unit == "B" else f"{value:.1f} {unit}"
            value /= 1024
    def set_busy(self, busy):
        state = "disabled" if busy else "normal"
        self.upload_button.configure(state=state); self.delete_button.configure(state=state); self.refresh_button.configure(state=state)
    def refresh(self):
        self.set_busy(True); self.status.set("Loading attachments…")
        def worker():
            try: result = self.service.list_attachments(); self.root.after(0, self.refresh_complete, result)
            except Exception as exc: self.root.after(0, self.failed, "Refresh", str(exc))
        threading.Thread(target=worker, daemon=True).start()
    def refresh_complete(self, attachments):
        for item in self.tree.get_children(): self.tree.delete(item)
        self.attachments_by_item.clear()
        for index, attachment in enumerate(attachments):
            item = self.tree.insert("", "end", values=(attachment["file_name"], self.display_size(attachment["file_size"]), attachment["upload_date"]))
            self.attachments_by_item[item] = attachment
        self.status.set(f"Loaded {len(attachments)} attachment(s)."); self.progress.configure(value=0); self.set_busy(False)
    def upload(self):
        files = filedialog.askopenfilenames(parent=self.root, title="Select attachment files")
        if not files: return
        self.set_busy(True); self.status.set("Uploading…"); self.progress.configure(value=0, maximum=100)
        def worker():
            outcomes = []
            for file_path in files:
                name = Path(file_path).name
                try:
                    self.service.upload(file_path, lambda sent,total,n=name:self.root.after(0, self.upload_progress, n, sent, total))
                    outcomes.append((name, True, "Upload Successful"))
                except Exception as exc: outcomes.append((name, False, f"Upload Failed: {exc}"))
            self.root.after(0, self.upload_complete, outcomes)
        threading.Thread(target=worker, daemon=True).start()
    def upload_progress(self, name, sent, total):
        self.status.set(f"Uploading... {name}"); self.progress.configure(value=(sent / total * 100) if total else 100)
    def upload_complete(self, outcomes):
        text = "\n".join(f"{name}: {message}" for name, _ok, message in outcomes); self.status.set(text)
        if all(ok for _name, ok, _message in outcomes): messagebox.showinfo(APP_NAME, text, parent=self.root)
        else: messagebox.showwarning(APP_NAME, text, parent=self.root)
        self.set_busy(False); self.refresh()
    def delete(self):
        selected = self.tree.selection()
        if not selected: messagebox.showwarning(APP_NAME, "Select an attachment to delete.", parent=self.root); return
        attachment = self.attachments_by_item[selected[0]]
        if not messagebox.askyesno(APP_NAME, f"Delete {attachment['file_name']}?", parent=self.root): return
        self.set_busy(True); self.status.set("Deleting...")
        def worker():
            try: self.service.delete(attachment["id"]); self.root.after(0, self.delete_complete)
            except Exception as exc: self.root.after(0, self.failed, "Delete", str(exc))
        threading.Thread(target=worker, daemon=True).start()
    def delete_complete(self):
        self.status.set("Deleted Successfully"); messagebox.showinfo(APP_NAME, "Deleted Successfully", parent=self.root); self.set_busy(False); self.refresh()
    def failed(self, operation, error):
        self.set_busy(False); self.progress.configure(value=0); self.status.set(f"{operation} failed: {error}"); messagebox.showerror(APP_NAME, f"The attachment server is unavailable or returned an error.\n\n{error}", parent=self.root)


class DatabaseReportsWindow:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent)
        self.root.title("Reports (Database)")
        self.root.geometry("520x220")
        self.root.resizable(False, False)
        frame = ttk.Frame(self.root, padding=24)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Reports (Database)", font=("Segoe UI", 17, "bold")).pack(anchor="w", pady=(0, 18))
        ttk.Label(frame, text="Database reporting will be available in a future version.", wraplength=440).pack(anchor="w")
        ttk.Button(frame, text="Close", command=self.root.destroy).pack(anchor="e", pady=(28, 0))


class TrackingSynchronizationWindow:
    def __init__(self, parent, on_config_saved=None, on_sync_complete=None):
        self.root = tk.Toplevel(parent)
        self.root.title("Tracking Synchronization")
        self.root.geometry("570x410")
        self.root.resizable(False, False)
        self.on_config_saved = on_config_saved
        self.on_sync_complete_callback = on_sync_complete
        self.cfg = load_config()
        frame = ttk.Frame(self.root, padding=22)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Tracking Synchronization", font=("Segoe UI", 17, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 20))
        self.enabled = tk.BooleanVar(value=self.cfg.get("tracking_sync_enabled", False))
        ttk.Checkbutton(frame, text="Enable Automatic Synchronization", variable=self.enabled).grid(row=1, column=0, columnspan=3, sticky="w", pady=8)
        ttk.Label(frame, text="Every").grid(row=2, column=0, sticky="w", pady=10)
        self.interval = tk.StringVar(value=str(self.cfg.get("tracking_sync_interval_hours", 5)))
        ttk.Spinbox(frame, from_=1, to=24, textvariable=self.interval, width=8).grid(row=2, column=1, sticky="w")
        ttk.Label(frame, text="Hours").grid(row=2, column=2, sticky="w")
        self.last_sync = tk.StringVar(value=self.cfg.get("tracking_last_sync_time") or "Never")
        ttk.Label(frame, text="Last Synchronization:").grid(row=3, column=0, sticky="nw", pady=10)
        ttk.Label(frame, textvariable=self.last_sync, wraplength=330).grid(row=3, column=1, columnspan=2, sticky="w", pady=10)
        self.status = tk.StringVar(value="Ready")
        ttk.Label(frame, textvariable=self.status, wraplength=500).grid(row=4, column=0, columnspan=3, sticky="w", pady=12)
        buttons = ttk.Frame(frame)
        buttons.grid(row=5, column=0, columnspan=3, sticky="w", pady=(18, 0))
        self.sync_button = ttk.Button(buttons, text="Sync Now", command=self.sync_now)
        self.sync_button.pack(side="left", padx=(0, 8))
        ttk.Button(buttons, text="Save Configuration", command=self.save_configuration).pack(side="left", padx=8)
        ttk.Button(buttons, text="Close", command=self.root.destroy).pack(side="left", padx=8)

    def save_configuration(self, show_confirmation=True):
        try:
            interval = int(self.interval.get())
            if interval < 1 or interval > 24:
                raise ValueError("Synchronization interval must be between 1 and 24 hours.")
            self.cfg["tracking_sync_enabled"] = self.enabled.get()
            self.cfg["tracking_sync_interval_hours"] = interval
            save_config(self.cfg)
            if self.on_config_saved:
                self.on_config_saved()
            if show_confirmation:
                messagebox.showinfo(APP_NAME, "Synchronization configuration saved.", parent=self.root)
        except ValueError as exc:
            messagebox.showerror(APP_NAME, str(exc), parent=self.root)

    def sync_now(self):
        self.save_configuration(show_confirmation=False)
        self.sync_button.configure(state="disabled")
        self.status.set("Synchronization started…")

        def worker():
            try:
                debug_mode = self.cfg.get("FullSynchronizationDebug", True)
                result = synchronize_with_bounce(self.cfg.get("tracking_last_sync_time", ""), debug_mode)
                if not debug_mode:
                    self.cfg["tracking_last_sync_time"] = result["last_sync_time"]
                save_config(self.cfg)
                self.root.after(0, self.sync_complete, result)
            except Exception as exc:
                self.root.after(0, self.sync_failed, str(exc))
        threading.Thread(target=worker, daemon=True).start()

    def sync_complete(self, result):
        self.sync_button.configure(state="normal")
        self.last_sync.set(result["last_sync_time"])
        text = ("Synchronization Complete\n\n"
                f"Records Downloaded: {result['records_downloaded']}\n"
                f"Rows Updated: {result['rows_updated']}\n"
                f"Bounces Detected: {result.get('bounce_detected', 0)}\n"
                f"Bounces Registered: {result.get('bounce_registered', 0)}\n"
                f"Bounce Rows Updated: {result.get('bounce_rows_updated', 0)}\n"
                f"Execution Time: {result['execution_time']:.2f} seconds\n"
                f"Last Synchronization Time: {result['last_sync_time']}")
        self.status.set(text)
        if self.on_sync_complete_callback:
            self.on_sync_complete_callback()
        messagebox.showinfo(APP_NAME, text, parent=self.root)

    def sync_failed(self, error):
        self.sync_button.configure(state="normal")
        self.status.set(f"Synchronization failed: {error}")
        messagebox.showerror(APP_NAME, f"Tracking synchronization failed:\n\n{error}", parent=self.root)


class EmailAutomationApp:
    def __init__(self):
        self.root=tk.Tk();self.root.title("Email Automation");self.root.geometry("1120x680");self.root.minsize(900,580)
        sidebar=tk.Frame(self.root,bg="#16324f",width=210);sidebar.pack(side="left",fill="y");sidebar.pack_propagate(False)
        tk.Label(sidebar,text="Email Automation",bg="#16324f",fg="white",font=("Segoe UI",16,"bold"),pady=24).pack(fill="x")
        actions=(("Dashboard",self.refresh),("Send Mail Now",lambda:SenderWindow(self.root)),("Daily Scheduling",lambda:MultiScheduleWindow(self.root)),("Mail Account Setup",lambda:AccountWindow(self.root)),("Tracking Synchronization",lambda:TrackingSynchronizationWindow(self.root,self.schedule_tracking_check,self.record_tracking_sync_run)),("Reports (Database)",lambda:DatabaseReportsWindow(self.root)),("Settings",lambda:SettingsWindow(self.root)),("Exit",self.root.destroy))
        for text,cmd in actions:tk.Button(sidebar,text=text,command=cmd,bg="#16324f",fg="white",activebackground="#285b87",activeforeground="white",relief="flat",anchor="w",padx=22,pady=11,font=("Segoe UI",10)).pack(fill="x")
        tk.Label(sidebar,text=f"Build:\n{BUILD_TIMESTAMP_UTC} UTC",bg="#16324f",fg="#b9cfe3",font=("Segoe UI",8),justify="left",padx=22,pady=12).pack(side="bottom",fill="x",anchor="w")
        self.main=ttk.Frame(self.root,padding=22);self.main.pack(side="left",fill="both",expand=True);self.sync_running=False;self.tracking_check_job=None;self.tracking_debug_last_run="";self.reply_check_running=False;self.reply_last_run=0.0;self.dashboard_page=1;self.dashboard_sort_key="sender";self.dashboard_sort_reverse=False;self.dashboard_search=tk.StringVar();self.refresh();self.tracking_check_job=self.root.after(1500,self.check_tracking_synchronization);self.root.after(5000,self.check_automatic_replies)

    def record_reply_check_run(self): self.reply_last_run = time.monotonic()

    def check_automatic_replies(self):
        interval = 5 * 60
        due = not self.reply_last_run or time.monotonic() - self.reply_last_run >= interval
        if due and not self.reply_check_running:
            self.reply_check_running = True
            def worker():
                try:
                    accounts = account_service(); service = reply_tracking_service()
                    for account in accounts.list_accounts():
                        configuration = accounts.imap_configuration(account["email"]) if account.get("enabled", True) else None
                        if configuration:
                            try: service.check_account(configuration)
                            except Exception: pass
                finally:
                    self.record_reply_check_run(); self.reply_check_running = False
            threading.Thread(target=worker, daemon=True).start()
        self.root.after(60000, self.check_automatic_replies)

    def record_tracking_sync_run(self):
        self.tracking_debug_last_run = datetime.now(timezone.utc).isoformat()

    def schedule_tracking_check(self):
        if self.tracking_check_job:
            self.root.after_cancel(self.tracking_check_job)
        self.tracking_check_job = self.root.after(100, self.check_tracking_synchronization)

    def check_tracking_synchronization(self):
        cfg = load_config()
        debug_mode = cfg.get("FullSynchronizationDebug", True)
        scheduling_cursor = self.tracking_debug_last_run if debug_mode else cfg.get("tracking_last_sync_time", "")
        due = synchronization_is_due(cfg.get("tracking_sync_enabled", False), cfg.get("tracking_sync_interval_hours", 5), scheduling_cursor)
        if due and not self.sync_running:
            self.sync_running = True
            last_sync = cfg.get("tracking_last_sync_time", "")
            def worker():
                try:
                    current = load_config(); debug_mode = current.get("FullSynchronizationDebug", True)
                    result = synchronize_with_bounce(last_sync, debug_mode)
                    if not debug_mode:
                        updated = load_config(); updated["tracking_last_sync_time"] = result["last_sync_time"]; save_config(updated)
                    else:
                        self.record_tracking_sync_run()
                except Exception:
                    pass
                finally:
                    self.sync_running = False
            threading.Thread(target=worker, daemon=True).start()
        self.tracking_check_job = self.root.after(60000, self.check_tracking_synchronization)
    def dashboard_sort_by(self,key):
        if self.dashboard_sort_key==key:self.dashboard_sort_reverse=not self.dashboard_sort_reverse
        else:self.dashboard_sort_key=key;self.dashboard_sort_reverse=False
        self.dashboard_page=1;self.refresh()
    def dashboard_goto_page(self,page):
        try:self.dashboard_page=int(page)
        except Exception:self.dashboard_page=1
        self.refresh()
    def dashboard_clear_search(self):
        self.dashboard_search.set("");self.dashboard_page=1;self.refresh()
    def refresh(self):
        for child in self.main.winfo_children():child.destroy()
        ttk.Label(self.main,text="Dashboard",font=("Segoe UI",20,"bold")).pack(anchor="w")
        accounts=account_service().list_accounts();summary=dashboard_excel_summary(page=self.dashboard_page,page_size=PAGE_SIZE,search=self.dashboard_search.get(),sort_key=self.dashboard_sort_key,sort_reverse=self.dashboard_sort_reverse);counts=summary["counts"];enabled=sum(1 for a in accounts if a.get("enabled",True));today_sent=summary["today_sent"];today_failed=summary["today_failed"];weekly_sent=summary["weekly_sent"];monthly_sent=summary["monthly_sent"]
        total_done=counts["sent"]+counts["failed"];success=(counts["sent"]/total_done*100) if total_done else 0
        metrics=(("Total Gmail Accounts",len(accounts)),("Enabled Gmail Accounts",enabled),("Disabled Gmail Accounts",len(accounts)-enabled),("Grand Total Emails",counts["total"]),("Pending Emails",counts["pending"]),("Sent Emails",counts["sent"]),("Failed Emails",counts["failed"]),("Today's Sent",today_sent),("Today's Failed",today_failed),("Weekly Sent",weekly_sent),("Monthly Sent",monthly_sent),("Success Rate",f"{success:.1f}%"),("Failure Rate",f"{100-success:.1f}%"))
        cards=ttk.Frame(self.main);cards.pack(fill="x",pady=18)
        for i,(label,value) in enumerate(metrics):
            box=ttk.LabelFrame(cards,text=label,padding=12);box.grid(row=i//4,column=i%4,sticky="nsew",padx=5,pady=5);ttk.Label(box,text=str(value),font=("Segoe UI",16,"bold")).pack();cards.columnconfigure(i%4,weight=1)
        ttk.Label(self.main,text="Per Sender Email Statistics",font=("Segoe UI",13,"bold")).pack(anchor="w",pady=(12,5))
        bar=ttk.Frame(self.main);bar.pack(fill="x",pady=(0,6));ttk.Label(bar,text="Search:").pack(side="left");ttk.Entry(bar,textvariable=self.dashboard_search,width=30).pack(side="left",padx=7);ttk.Button(bar,text="Search",command=lambda:(setattr(self,"dashboard_page",1),self.refresh())).pack(side="left");ttk.Button(bar,text="Clear",command=self.dashboard_clear_search).pack(side="left",padx=7);ttk.Button(bar,text="Refresh",command=self.refresh).pack(side="right")
        tree=ttk.Treeview(self.main,columns=("sender","sent","pending"),show="headings",height=8)
        for k,t in (("sender","Sender Email"),("sent","Sent Today"),("pending","Pending")):tree.heading(k,text=t,command=lambda value=k:self.dashboard_sort_by(value))
        for row in summary["per_rows"]:tree.insert("","end",values=row)
        tree.pack(fill="both",expand=True)
        paging=ttk.Frame(self.main);paging.pack(fill="x",pady=8);ttk.Button(paging,text="First",command=lambda:self.dashboard_goto_page(1)).pack(side="left");ttk.Button(paging,text="Previous",command=lambda:self.dashboard_goto_page(self.dashboard_page-1)).pack(side="left",padx=4);ttk.Label(paging,text="Page").pack(side="left",padx=(10,3));page_number=tk.StringVar(value=str(summary["per_page"]));entry=ttk.Entry(paging,textvariable=page_number,width=5);entry.pack(side="left");entry.bind("<Return>",lambda _event:self.dashboard_goto_page(page_number.get()));ttk.Label(paging,text=f"of {summary['per_total_pages']}  ({summary['per_filtered_total']} rows)").pack(side="left",padx=4);ttk.Button(paging,text="Next",command=lambda:self.dashboard_goto_page(summary["per_page"]+1)).pack(side="left",padx=(10,4));ttk.Button(paging,text="Last",command=lambda:self.dashboard_goto_page(summary["per_total_pages"])).pack(side="left")
        self.dashboard_page=summary["per_page"]
    def run(self):self.root.mainloop()


def verify_app():
    root = tk.Tk(); root.withdraw(); paths = app_paths(); desktop = Path.home() / "Desktop"
    checks = []
    try: credentials(); checks.append(("SMTP configuration", True, "Credentials found"))
    except Exception as exc: checks.append(("SMTP configuration", False, str(exc)))
    checks.append(("Excel file", paths["list"].exists(), str(paths["list"])))
    shortcuts = [desktop / "Send Pending Emails Now.lnk", desktop / "Configure Daily Email Schedule.lnk", desktop / "Verify Installation.lnk"]
    checks.append(("Desktop shortcuts", all(x.exists() for x in shortcuts), str(desktop)))
    task = task_query(); checks.append(("Scheduled task", task["exists"], "Enabled" if task["enabled"] else "Disabled or missing"))
    checks.append(("Config files", config_path().exists() and paths["env"].exists(), str(config_path())))
    text = "INSTALLATION VERIFICATION\n\n" + "\n".join(f"{'PASS' if ok else 'FAIL'}  {name}\n      {detail}" for name, ok, detail in checks)
    ensure_dirs()
    messagebox.showinfo("Verify Installation", text); root.destroy()


def sample_workbook(path):
    wb = Workbook(); ws = wb.active; ws.title = "Mail List"
    headers = ["First_Name", "Last_Name", "Email", "Company", "Designation", "Country", "Subject", "Body", "Sender_Name", "Sender_Email", "Status", "Result", "SentDate"]
    ws.append(headers)
    ws.append(["John", "Doe", "recipient@example.com", "Example Company", "Manager", "USA", "Opportunity for {{Company}}", "Hi {{First_Name}} {{Last_Name}},\n\nA message for {{Company}}.", "Power Soft", "sales@gmail.com", "Pending", "", ""])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = "A1:M2"
    widths = (16, 16, 30, 24, 24, 16, 36, 60, 22, 30, 14, 40, 22)
    for column, width in enumerate(widths, 1): ws.column_dimensions[ws.cell(1, column).column_letter].width = width
    wb.save(path)


def main():
    name = Path(sys.executable if getattr(sys, "frozen", False) else sys.argv[0]).stem.lower()
    if "--run-schedule" in sys.argv:
        try:
            schedule_id=sys.argv[sys.argv.index("--run-schedule")+1];item=next(x for x in load_schedules() if x["id"]==schedule_id and x.get("enabled",True));send_pending(int(item["max_emails"]),True,excel_path=item["excel_file"],attachment_ids=item.get("attachment_ids",[]),attachment_source=item.get("attachment_source","Attachment Library"),local_attachment_paths=item.get("local_attachment_paths",[]))
        except Exception: pass
        return
    if name == "email automation" or name == "emailautomation":
        migrate_legacy_account();EmailAutomationApp().run();return
    if "setup" in name: setup_app()
    elif "configure" in name: ScheduleWindow().run()
    elif "verify" in name: verify_app()
    elif "--scheduled" in sys.argv:
        try: send_pending(int(load_config()["daily_limit"]), True)
        except Exception: pass
    else: manual_send()


if __name__ == "__main__":
    main()


