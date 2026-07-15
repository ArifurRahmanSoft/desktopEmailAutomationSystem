import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
from urllib.error import URLError

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import email_automation as app


class FakeSMTP:
    messages = []

    def send_message(self, message):
        self.messages.append(message)

    def quit(self):
        pass


class FakeAccounts:
    def smtp_configuration(self, email):
        if email != "sender@example.com":
            return None
        return {"email": email, "password": "app-password", "smtp_host": "smtp.test", "smtp_port": 587, "encryption": "STARTTLS"}

    def connect_smtp(self, configuration, timeout=45):
        return FakeSMTP()


class FakeResponse:
    status = 200

    def __enter__(self):
        return self

    def __exit__(self, *args):
        pass

    def read(self):
        return b'{"success": true}'


def write_settings(installed, data):
    (installed / "config").mkdir(parents=True)
    (installed / "config" / "settings.json").write_text(json.dumps({"data_dir": str(data), "backup_enabled": False, "random_delay_min": 0, "random_delay_max": 0}), encoding="utf-8")


def create_mail_list(path, total_rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["First_Name", "Last_Name", "Email", "Company", "Project_Name", "Subject", "Body", "Sender_Email", "Status", "Result", "SentDate"])
    for index in range(1, total_rows + 1):
        status = "Pending" if index % 3 else "Sent"
        ws.append([f"First{index:04d}", f"Last{index:04d}", f"user{index:04d}@example.com", "PowerSoft" if index % 2 else "Other", "Project A", f"Subject {index:04d}", "Body", "sender@example.com", status, "", ""])
    wb.save(path)


class PaginationAndSendRegistrationTests(unittest.TestCase):
    def test_email_grid_page_returns_exactly_twenty_rows(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "mail_list.xlsx"
            create_mail_list(path, 45)
            first = app.email_grid_page(path, page=1)
            last = app.email_grid_page(path, page=3)
            self.assertEqual(len(first["rows"]), 20)
            self.assertEqual(first["page"], 1)
            self.assertEqual(first["total_pages"], 3)
            self.assertEqual(len(last["rows"]), 5)

    def test_email_grid_search_and_sort_use_filtered_pagination(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "mail_list.xlsx"
            create_mail_list(path, 45)
            page = app.email_grid_page(path, page=1, search="user", sort_key="email", sort_reverse=True)
            self.assertEqual(len(page["rows"]), 20)
            self.assertEqual(page["filtered_total"], 45)
            emails = [row[1] for row in page["rows"]]
            self.assertEqual(emails, sorted(emails, reverse=True))

    def test_send_registration_called_once_after_each_successful_send(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            data = root / "data"
            data.mkdir()
            installed = root / "installed"
            write_settings(installed, data)
            wb = Workbook()
            ws = wb.active
            ws.append(["Email", "Project_Name", "Subject", "Body", "Sender_Email", "Status", "Result", "SentDate"])
            ws.append(["one@example.com", "Alpha", "Subject One", "Body", "sender@example.com", "Pending", "", ""])
            ws.append(["two@example.com", "Beta", "Subject Two", "Body", "sender@example.com", "Pending", "", ""])
            wb.save(data / "mail_list.xlsx")
            requests = []

            def fake_urlopen(request, timeout=20):
                requests.append((request.get_method(), request.full_url, json.loads(request.data.decode("utf-8"))))
                return FakeResponse()

            FakeSMTP.messages = []
            with patch.object(app, "install_dir", return_value=installed), patch.object(app, "account_service", return_value=FakeAccounts()), patch.object(app, "domain_has_mx", return_value=True), patch.object(app, "urlopen", fake_urlopen):
                result = app.send_pending(2, wait_between=False)
            self.assertEqual(result["sent"], 2)
            self.assertEqual(result["failed"], 0)
            self.assertEqual(len(FakeSMTP.messages), 2)
            self.assertEqual(len(requests), 2)
            self.assertTrue(all(method == "POST" for method, _url, _payload in requests))
            self.assertTrue(all(url == "https://emailtrackingserver-v2-2.onrender.com/api/tracking/register-send" for _method, url, _payload in requests))
            payloads = [payload for _method, _url, payload in requests]
            self.assertEqual([item["recipient_mail"] for item in payloads], ["one@example.com", "two@example.com"])
            self.assertEqual([item["project_name"] for item in payloads], ["Alpha", "Beta"])
            self.assertTrue(all(item["tracking_id"] for item in payloads))
            self.assertTrue(all(item["message_id"] for item in payloads))
            self.assertEqual([message["Message-ID"] for message in FakeSMTP.messages], [item["message_id"] for item in payloads])
            self.assertTrue(all(item["message_id"].startswith("<") and item["message_id"].endswith("@emailautomation-v2.local>") for item in payloads))
            self.assertTrue(all(item["sent_time"] for item in payloads))

    def test_failed_send_registration_does_not_stop_successful_send(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            data = root / "data"
            data.mkdir()
            installed = root / "installed"
            write_settings(installed, data)
            wb = Workbook()
            ws = wb.active
            ws.append(["Email", "Project_Name", "Subject", "Body", "Sender_Email", "Status", "Result", "SentDate"])
            ws.append(["one@example.com", "Alpha", "Subject One", "Body", "sender@example.com", "Pending", "", ""])
            wb.save(data / "mail_list.xlsx")

            def failing_urlopen(*_args, **_kwargs):
                raise RuntimeError("registration API offline")

            FakeSMTP.messages = []
            with patch.object(app, "install_dir", return_value=installed), patch.object(app, "account_service", return_value=FakeAccounts()), patch.object(app, "domain_has_mx", return_value=True), patch.object(app, "urlopen", failing_urlopen):
                result = app.send_pending(1, wait_between=False)
            self.assertEqual(result["sent"], 1)
            self.assertEqual(result["failed"], 0)
            check = load_workbook(data / "mail_list.xlsx", data_only=True)
            sheet = check.active
            headers = app.header_map(sheet)
            self.assertEqual(sheet.cell(2, headers["status"]).value, "Sent")
            self.assertEqual(sheet.cell(2, headers["result"]).value, "Success")
            debug_text = (installed / "debug" / "tracking-registration-debug.log").read_text(encoding="utf-8")
            self.assertIn("registration API offline", debug_text)
            self.assertIn("Registration Failed", debug_text)
            check.close()

    def test_network_registration_failure_retries_once_without_resending_email(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            data = root / "data"
            data.mkdir()
            installed = root / "installed"
            write_settings(installed, data)
            wb = Workbook()
            ws = wb.active
            ws.append(["Email", "Project_Name", "Subject", "Body", "Sender_Email", "Status", "Result", "SentDate"])
            ws.append(["one@example.com", "Alpha", "Subject One", "Body", "sender@example.com", "Pending", "", ""])
            wb.save(data / "mail_list.xlsx")
            calls = []

            def flaky_urlopen(request, timeout=20):
                calls.append(request.full_url)
                raise URLError("temporary outage")

            FakeSMTP.messages = []
            with patch.object(app, "install_dir", return_value=installed), patch.object(app, "account_service", return_value=FakeAccounts()), patch.object(app, "domain_has_mx", return_value=True), patch.object(app, "urlopen", flaky_urlopen), patch.object(app.time, "sleep") as sleep:
                result = app.send_pending(1, wait_between=False)
            self.assertEqual(result["sent"], 1)
            self.assertEqual(result["failed"], 0)
            self.assertEqual(len(FakeSMTP.messages), 1)
            self.assertEqual(len(calls), 2)
            sleep.assert_called_once_with(2)
            debug_text = (installed / "debug" / "tracking-registration-debug.log").read_text(encoding="utf-8")
            self.assertIn("Registration Network Error", debug_text)
            self.assertIn("Tracking registration network error after retry", debug_text)


if __name__ == "__main__":
    unittest.main()

