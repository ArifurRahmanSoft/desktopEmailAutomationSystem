import tempfile
import sys
import re
import json
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import email_automation as app


class FakeSMTP:
    messages = []
    connections = 0
    def __init__(self, *args, **kwargs): self.__class__.connections += 1
    def __enter__(self): return self
    def __exit__(self, *args): pass
    def ehlo(self): pass
    def starttls(self): pass
    def login(self, *args): pass
    def send_message(self, msg): self.messages.append(msg)
    def quit(self): pass


class FakeAccounts:
    def smtp_configuration(self, email):
        return {"email": email, "sender_alias": "alias@powersoft.info", "provider": "Gmail", "password": "app-password", "smtp_host": "smtp.gmail.com", "smtp_port": 587, "encryption": "STARTTLS"} if email == "sender@example.com" else None
    def connect_smtp(self, configuration, timeout=45):
        return FakeSMTP()


class FakeResponse:
    status = 200
    def __enter__(self): return self
    def __exit__(self, *args): pass
    def read(self): return b'{"success": true}'


registration_requests = []


def fake_urlopen(request, timeout=20):
    registration_requests.append((request.get_method(), request.full_url, json.loads(request.data.decode("utf-8"))))
    return FakeResponse()


with tempfile.TemporaryDirectory() as temp:
    root = Path(temp)
    data = root / "data"
    data.mkdir()
    installed = root / "installed"
    (installed / "config").mkdir(parents=True)
    (installed / ".env").write_text("EMAIL_ADDRESS=test@example.com\nEMAIL_PASSWORD=app-password\n", encoding="utf-8")
    (installed / "config" / "settings.json").write_text(
        '{"daily_limit": 5, "schedule_time": "09:00", "default_sender_name": "Default Sender", "data_dir": "' + str(data).replace("\\", "\\\\") + '"}',
        encoding="utf-8",
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["First_Name", "Last_Name", "Email", "Company", "Project_Name", "Future_Field", "Subject", "Body", "Sender_Name", "Sender_Email", "Status", "Result", "SentDate"])
    ws.append(["Valid", "Person", "valid@example.com", None, "PowerSoft", "Future Value", "Hello {{ FIRST_NAME }} {{unknown}}", "{{last_name}} at {{Company}} / {{future_field}}\nVisit https://powersoft.com", "Row Sender", "sender@example.com", "Pending", "", ""])
    ws.append(["Invalid", "Person", "invalid-address", "Example", "PowerSoft", "", "Bad", "Body", "", "sender@example.com", "Pending", "", ""])
    wb.save(data / "mail_list.xlsx")
    with patch.object(app, "install_dir", return_value=installed), patch.object(app.smtplib, "SMTP", FakeSMTP), patch.object(app, "account_service", return_value=FakeAccounts()), patch.object(app, "domain_has_mx", return_value=True), patch.object(app, "urlopen", fake_urlopen):
        result = app.send_pending(5, wait_between=False)
        assert result == {"requested": 5, "sent": 1, "failed": 1, "remaining": 0}, result
        result_book = load_workbook(data / "mail_list.xlsx", data_only=True)
        out = result_book.active
        headers = app.header_map(out)
        assert out.cell(2, headers["status"]).value == "Sent"
        assert out.cell(2, headers["result"]).value == "Success"
        assert out.cell(3, headers["status"]).value == "Failed"
        assert "Invalid email" in out.cell(3, headers["result"]).value
        assert out.cell(3, headers["sentdate"]).value is not None
        assert len(FakeSMTP.messages) == 1
        assert FakeSMTP.connections == 1
        assert FakeSMTP.messages[0]["Subject"] == "Hello Valid"
        assert FakeSMTP.messages[0]["From"].startswith("Row Sender <")
        assert FakeSMTP.messages[0]["From"] == "Row Sender <alias@powersoft.info>"
        plain_part = FakeSMTP.messages[0].get_body(preferencelist=("plain",))
        html_part = FakeSMTP.messages[0].get_body(preferencelist=("html",))
        assert "Person at  / Future Value\nVisit https://powersoft.com" in plain_part.get_content()
        tracking_id = out.cell(2, headers["trackingid"]).value
        assert len(registration_requests) == 1
        method, url, payload = registration_requests[0]
        assert method == "POST"
        assert url == "https://emailtrackingserver-v2-2.onrender.com/api/tracking/register-send"
        assert payload["tracking_id"] == tracking_id
        assert payload["sender_mail"] == "sender@example.com"
        assert payload["recipient_mail"] == "valid@example.com"
        assert payload["mail_subject"] == "Hello Valid"
        assert payload["project_name"] == "PowerSoft"
        assert payload["excel_file_path"] == str((data / "mail_list.xlsx").resolve())
        assert payload["sent_time"]
        assert payload["message_id"] == FakeSMTP.messages[0]["Message-ID"]
        assert re.fullmatch(r"<[0-9a-f]{32}\.[0-9]{20}@emailautomation-v2\.local>", payload["message_id"])
        assert re.fullmatch(r"[0-9a-f]{8}-[0-9a-f]{4}-4[0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}", tracking_id)
        assert plain_part.get_content().strip() == "Person at  / Future Value\nVisit https://powersoft.com"
        assert f'https://emailtrackingserver-v2-2.onrender.com/email/open/{tracking_id}' in html_part.get_content()
        assert '<img src="https://emailtrackingserver-v2-2.onrender.com/email/open/' in html_part.get_content()
        click_url = f'https://emailtrackingserver-v2-2.onrender.com/email/click/{tracking_id}?url=https%3A%2F%2Fpowersoft.com'
        assert click_url in html_part.get_content()
        assert '>https://powersoft.com</a>' in html_part.get_content()
        assert len(list((installed / "backup").glob("*.xlsx"))) == 1
        assert len(list((installed / "logs").glob("*.log"))) == 0
        result_book.close()
print("SMOKE TEST PASS")

