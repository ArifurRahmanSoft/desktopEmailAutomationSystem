import sys
import tempfile
import unittest
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from bounce_tracking_service import BounceTrackingService, extract_bounce_details, is_bounce_message


def bounce_message():
    message = EmailMessage()
    message["From"] = "Mail Delivery Subsystem <mailer-daemon@example.com>"
    message["Subject"] = "Delivery Status Notification (Failure)"
    message["Message-ID"] = "<bounce-1@example.com>"
    message.set_content("Final-Recipient: rfc822; failed@example.com\nDiagnostic-Code: smtp; 550 5.1.1 User unknown")
    return message


class FakeIMAP:
    def __init__(self, messages): self.messages = messages; self.seen = []; self.logged_out = False
    def login(self, email, password): self.login_values = (email, password)
    def select(self, folder): return "OK", [b""]
    def search(self, *_): return "OK", [b"1 2"]
    def fetch(self, number, _): return "OK", [(b"RFC822", self.messages[int(number)-1])]
    def store(self, number, *_): self.seen.append(number); return "OK", []
    def noop(self): return "OK", []
    def logout(self): self.logged_out = True


class BounceTrackingTests(unittest.TestCase):
    def workbook(self, folder):
        path = Path(folder) / "mail_list.xlsx"; wb = Workbook(); ws = wb.active
        ws.append(["Email", "Status", "Result", "Unrelated"]); ws.append(["failed@example.com", "Sent", "Success", "keep"]); ws.append(["other@example.com", "Sent", "Success", "untouched"]); wb.save(path); return path

    def test_detects_and_extracts_common_bounce(self):
        message = bounce_message(); self.assertTrue(is_bounce_message(message)); recipient, reason = extract_bounce_details(message)
        self.assertEqual(recipient, "failed@example.com"); self.assertIn("550 5.1.1 User unknown", reason)

    def test_updates_only_one_matched_row_and_adds_columns(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp); service = BounceTrackingService(path, Path(temp) / "logs")
            row = service.update_excel("failed@example.com", "550 User unknown"); self.assertEqual(row, 2)
            wb = load_workbook(path, data_only=True); ws = wb.active; headers = {str(c.value):i for i,c in enumerate(ws[1],1)}
            self.assertEqual(ws.cell(2, headers["Status"]).value, "Failed"); self.assertEqual(ws.cell(2, headers["Result"]).value, "550 User unknown")
            self.assertEqual(ws.cell(2, headers["BounceStatus"]).value, "Yes"); self.assertEqual(ws.cell(2, headers["BounceReason"]).value, "550 User unknown"); self.assertTrue(ws.cell(2, headers["BounceTime"]).value)
            self.assertIsNone(ws.cell(3, headers["BounceStatus"]).value); self.assertEqual(ws.cell(3, headers["Status"]).value, "Sent"); self.assertEqual(ws.cell(3, headers["Result"]).value, "Success"); self.assertEqual(ws.cell(2, headers["Unrelated"]).value, "keep"); wb.close()

    def test_unread_duplicate_message_is_processed_once(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp); raw = bounce_message().as_bytes(); client = FakeIMAP([raw, raw])
            service = BounceTrackingService(path, Path(temp) / "logs", imap_ssl_factory=lambda *_: client)
            result = service.check_account({"email":"sender@example.com","password":"secret","imap_host":"imap.example.com","imap_port":993,"imap_encryption":"SSL/TLS"})
            self.assertEqual(result, {"detected":1,"matched":1}); self.assertEqual(len(client.seen),2); self.assertTrue(client.logged_out)
            log = next((Path(temp)/"logs").glob("bounce-*.log")).read_text(encoding="utf-8"); self.assertEqual(log.count('"message_id"'),1)

    def test_test_connection_does_not_read_or_send_mail(self):
        client = FakeIMAP([]); service = BounceTrackingService("unused.xlsx", ".", imap_ssl_factory=lambda *_: client)
        self.assertTrue(service.test_connection({"email":"sender@example.com","password":"secret","imap_host":"imap.example.com","imap_port":993,"imap_encryption":"SSL/TLS"})); self.assertTrue(client.logged_out)


if __name__ == "__main__": unittest.main()
