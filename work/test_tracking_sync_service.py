import hashlib
import sys
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from tracking_sync_service import TRACKING_COLUMNS, TrackingSynchronizationService, synchronization_is_due


class TrackingSynchronizationTests(unittest.TestCase):
    def workbook(self, folder):
        path = Path(folder) / "mail_list.xlsx"
        wb = Workbook(); ws = wb.active
        ws.append(["Email", "Status", "TrackingId", "Unrelated"])
        ws.append(["one@example.com", "Sent", "track-1", "keep-one"])
        ws.append(["two@example.com", "Sent", "track-2", "keep-two"])
        wb.save(path)
        return path

    def workbook_named(self, folder, name, rows):
        path = Path(folder) / name
        wb = Workbook(); ws = wb.active
        ws.append(["Email", "Status", "TrackingId", "Unrelated"])
        for row in rows:
            ws.append(row)
        wb.save(path)
        return path

    def test_first_sync_downloads_all_and_updates_only_tracking_columns(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp); urls = []
            payload = [
                {"excel_file_path": str(path), "trackingId": "track-1", "openCount": 3, "clickCount": 2, "firstOpen": "2026-01-01", "lastOpen": "2026-01-02", "firstClick": "2026-01-03", "lastClick": "2026-01-04", "last_synchronize_time": "2026-07-11T10:00:00Z"},
                {"excel_file_path": str(path), "tracking_id": "missing-row", "open_count": 99},
            ]
            service = TrackingSynchronizationService("https://server.test", path, Path(temp) / "logs", lambda url: urls.append(url) or payload, lambda *_: {})
            result = service.sync("")
            self.assertEqual(urls, ["https://server.test/api/tracking/sync"])
            self.assertEqual(result["records_downloaded"], 2)
            self.assertEqual(result["rows_updated"], 1)
            self.assertEqual(result["last_sync_time"], "")
            wb = load_workbook(path, data_only=True); ws = wb.active
            headers = {str(c.value): i for i, c in enumerate(ws[1], 1)}
            self.assertEqual(ws.cell(2, headers["OpenCount"]).value, 3)
            self.assertEqual(ws.cell(2, headers["ClickCount"]).value, 2)
            self.assertEqual(ws.cell(2, headers["LastSynchronizeTime"]).value, "2026-07-11T10:00:00Z")
            self.assertEqual(ws.cell(2, headers["Unrelated"]).value, "keep-one")
            self.assertEqual(ws.cell(2, headers["Status"]).value, "Sent")
            self.assertTrue(all(name in headers for name in TRACKING_COLUMNS))
            wb.close()

    def test_incremental_sync_uses_updated_after(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp); urls = []
            service = TrackingSynchronizationService("https://server.test", path, Path(temp) / "logs", lambda url: urls.append(url) or [], lambda *_: {})
            service.sync("2026-06-30T10:20:30+00:00")
            self.assertIn("updated_after=2026-06-30T10%3A20%3A30%2B00%3A00", urls[0])

    def test_download_fields_are_appended_and_updated_without_local_logs(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp)
            wb = load_workbook(path)
            ws = wb.active
            for name in ("OpenCount", "ClickCount", "FirstOpen", "LastOpen", "FirstClick", "LastClick"):
                ws.cell(1, ws.max_column + 1, name)
            ws.cell(2, 4, "keep-one")
            wb.save(path)
            wb.close()
            payload = [
                {
                    "tracking_id": "track-1",
                    "excel_file_path": str(path),
                    "download_count": 4,
                    "first_download": "2026-07-04T01:02:03Z",
                    "last_download": "2026-07-04T04:05:06Z",
                },
                {"excel_file_path": str(path), "trackingId": "missing-row", "downloadCount": 2},
                {"excel_file_path": str(path), "trackingId": "track-2", "openCount": 8},
            ]
            logs = Path(temp) / "logs"
            service = TrackingSynchronizationService("https://server.test", path, logs, lambda _: payload, lambda *_: {})
            result = service.sync("")
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            headers = {str(cell.value): index for index, cell in enumerate(ws[1], 1)}
            self.assertEqual(headers["DownloadCount"], headers["LastClick"] + 1)
            self.assertEqual(headers["FirstDownload"], headers["LastClick"] + 2)
            self.assertEqual(headers["LastDownload"], headers["LastClick"] + 3)
            self.assertEqual(ws.cell(2, headers["DownloadCount"]).value, 4)
            self.assertEqual(ws.cell(2, headers["FirstDownload"]).value, "2026-07-04T01:02:03Z")
            self.assertEqual(ws.cell(2, headers["LastDownload"]).value, "2026-07-04T04:05:06Z")
            self.assertEqual(ws.cell(2, headers["Unrelated"]).value, "keep-one")
            wb.close()
            self.assertEqual(result["download_records_received"], 2)
            self.assertEqual(result["download_rows_updated"], 1)
            self.assertFalse(list(logs.glob("*.log")))

    def test_cursor_uses_maximum_server_updated_at_without_local_logs(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp)
            payload = [
                {"trackingId": "track-1", "openCount": 1, "updated_at": "2026-06-30T09:00:00Z"},
                {"trackingId": "track-2", "clickCount": 2, "updatedAt": "2026-06-30T12:30:00+00:00"},
                {"trackingId": "track-2", "openCount": 3},
            ]
            logs = Path(temp) / "logs"
            service = TrackingSynchronizationService("https://server.test", path, logs, lambda _: payload, lambda *_: {})
            result = service.sync("2026-06-30T08:00:00Z")
            self.assertEqual(result["last_sync_time"], "2026-06-30T12:30:00+00:00")
            self.assertFalse(list(logs.glob("*.log")))

    def test_zero_records_keeps_previous_cursor(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp)
            previous = "2026-06-30T08:00:00Z"
            service = TrackingSynchronizationService("https://server.test", path, Path(temp) / "logs", lambda _: [], lambda *_: {})
            result = service.sync(previous)
            self.assertEqual(result["last_sync_time"], previous)

    def test_records_without_excel_file_path_do_not_update_fallback_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp)
            payload = [{"tracking_id": "track-1", "open_count": 55, "updated_at": "2026-07-11T12:00:00Z"}]
            service = TrackingSynchronizationService("https://server.test", path, Path(temp) / "logs", lambda _: payload, lambda *_: {})

            result = service.sync("")

            workbook = load_workbook(path, data_only=True); worksheet = workbook.active
            headers = {str(cell.value): index for index, cell in enumerate(worksheet[1], 1)}
            self.assertNotIn("OpenCount", headers)
            workbook.close()
            self.assertEqual(result["rows_updated"], 0)
            self.assertEqual(result["records_missing_excel_path"], 1)

    def test_full_debug_sync_ignores_cursor_and_returns_reconciliation_counts(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp); urls = []
            payload = [
                {"excel_file_path": str(path), "trackingId": "track-1", "openCount": 4, "updated_at": "2026-06-30T12:00:00Z"},
                {"excel_file_path": str(path), "trackingId": "not-in-excel", "clickCount": 1, "updated_at": "2026-06-30T13:00:00Z"},
            ]
            previous = "2026-06-29T08:00:00Z"; logs = Path(temp) / "logs"
            service = TrackingSynchronizationService("https://server.test", path, logs, lambda url: urls.append(url) or payload, lambda *_: {})
            result = service.sync(previous, full_synchronization_debug=True)
            self.assertEqual(urls, ["https://server.test/api/tracking/sync"])
            self.assertEqual(result["last_sync_time"], previous)
            self.assertEqual(result["tracking_ids_matched"], 1)
            self.assertEqual(result["rows_updated"], 1)
            self.assertEqual(result["tracking_ids_not_found"], 1)
            self.assertFalse(list(logs.glob("*.log")))

    def test_api_failure_does_not_modify_excel(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(temp)
            before = hashlib.sha256(path.read_bytes()).hexdigest()
            def failed(_): raise RuntimeError("API unavailable")
            service = TrackingSynchronizationService("https://server.test", path, Path(temp) / "logs", failed, lambda *_: {})
            with self.assertRaisesRegex(RuntimeError, "API unavailable"): service.sync("")
            self.assertEqual(before, hashlib.sha256(path.read_bytes()).hexdigest())

    def test_two_excel_files_are_synchronized_in_one_cycle(self):
        with tempfile.TemporaryDirectory() as temp:
            first = self.workbook_named(temp, "first.xlsx", [["same@example.com", "Sent", "track-a", "first"]])
            second = self.workbook_named(temp, "second.xlsx", [["same@example.com", "Sent", "track-b", "second"]])
            posts = []
            payload = [
                {"excel_file_path": str(first), "tracking_id": "track-a", "open_count": 5, "click_count": 1, "updated_at": "2026-07-09T10:00:00Z"},
                {"excel_file_path": str(second), "tracking_id": "track-b", "open_count": 7, "download_count": 2, "updated_at": "2026-07-09T10:01:00Z"},
            ]
            service = TrackingSynchronizationService(
                "https://server.test",
                first,
                Path(temp) / "logs",
                lambda _: payload,
                lambda url, body: posts.append((url, body)) or {},
            )

            result = service.sync("")

            first_book = load_workbook(first, data_only=True); first_sheet = first_book.active
            first_headers = {str(cell.value): index for index, cell in enumerate(first_sheet[1], 1)}
            second_book = load_workbook(second, data_only=True); second_sheet = second_book.active
            second_headers = {str(cell.value): index for index, cell in enumerate(second_sheet[1], 1)}
            self.assertEqual(first_sheet.cell(2, first_headers["OpenCount"]).value, 5)
            self.assertEqual(first_sheet.cell(2, first_headers["ClickCount"]).value, 1)
            self.assertEqual(second_sheet.cell(2, second_headers["OpenCount"]).value, 7)
            self.assertEqual(second_sheet.cell(2, second_headers["DownloadCount"]).value, 2)
            self.assertEqual(result["workbooks_processed"], 2)
            self.assertEqual(result["rows_updated"], 2)
            self.assertEqual(len(posts), 2)
            self.assertEqual({body["tracking_id"] for _, body in posts}, {"track-a", "track-b"})
            first_book.close(); second_book.close()

    def test_sync_matches_tracking_id_only_not_email_address(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook_named(
                temp,
                "mail_list.xlsx",
                [
                    ["shared@example.com", "Sent", "correct-id", "correct-row"],
                    ["shared@example.com", "Sent", "other-id", "other-row"],
                ],
            )
            payload = [
                {
                    "excel_file_path": str(path),
                    "tracking_id": "other-id",
                    "email": "shared@example.com",
                    "open_count": 9,
                }
            ]
            service = TrackingSynchronizationService("https://server.test", path, Path(temp) / "logs", lambda _: payload, lambda *_: {})

            service.sync("")

            workbook = load_workbook(path, data_only=True); worksheet = workbook.active
            headers = {str(cell.value): index for index, cell in enumerate(worksheet[1], 1)}
            self.assertIsNone(worksheet.cell(2, headers["OpenCount"]).value)
            self.assertEqual(worksheet.cell(3, headers["OpenCount"]).value, 9)
            workbook.close()

    def test_missing_workbook_does_not_stop_synchronization(self):
        with tempfile.TemporaryDirectory() as temp:
            existing = self.workbook_named(temp, "existing.xlsx", [["one@example.com", "Sent", "track-1", "keep"]])
            missing = Path(temp) / "missing.xlsx"
            payload = [
                {"excel_file_path": str(missing), "tracking_id": "missing-track", "open_count": 99},
                {"excel_file_path": str(existing), "tracking_id": "track-1", "open_count": 3},
            ]
            service = TrackingSynchronizationService("https://server.test", existing, Path(temp) / "logs", lambda _: payload, lambda *_: {})

            result = service.sync("")

            workbook = load_workbook(existing, data_only=True); worksheet = workbook.active
            headers = {str(cell.value): index for index, cell in enumerate(worksheet[1], 1)}
            self.assertEqual(worksheet.cell(2, headers["OpenCount"]).value, 3)
            workbook.close()
            self.assertEqual(result["missing_workbooks"], 1)
            self.assertEqual(result["rows_updated"], 1)

    def test_bounce_fields_update_correct_workbook_by_tracking_id(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook_named(temp, "bounce.xlsx", [["bad@example.com", "Sent", "bounce-track", "keep"]])
            payload = [
                {
                    "excel_file_path": str(path),
                    "tracking_id": "bounce-track",
                    "is_bounce": True,
                    "bounce_time": "2026-07-13T10:30:00Z",
                    "bounce_reason": "550 5.1.1 User unknown",
                }
            ]
            service = TrackingSynchronizationService("https://server.test", None, Path(temp) / "logs", lambda _: payload, lambda *_: {})

            result = service.sync("")

            workbook = load_workbook(path, data_only=True); worksheet = workbook.active
            headers = {str(cell.value): index for index, cell in enumerate(worksheet[1], 1)}
            self.assertEqual(worksheet.cell(2, headers["Status"]).value, "Bounce")
            self.assertEqual(worksheet.cell(2, headers["is_bounce"]).value, True)
            self.assertEqual(worksheet.cell(2, headers["bounce_time"]).value, "2026-07-13T10:30:00Z")
            self.assertEqual(worksheet.cell(2, headers["bounce_reason"]).value, "550 5.1.1 User unknown")
            self.assertEqual(worksheet.cell(2, headers["Unrelated"]).value, "keep")
            workbook.close()
            self.assertEqual(result["bounce_records_received"], 1)
            self.assertEqual(result["bounce_rows_updated"], 1)

    def test_schedule_due_supports_first_run_interval_and_restart_catchup(self):
        now = datetime(2026, 6, 30, 12, 0, tzinfo=timezone.utc)
        self.assertTrue(synchronization_is_due(True, 5, "", now))
        self.assertFalse(synchronization_is_due(False, 5, "", now))
        self.assertFalse(synchronization_is_due(True, 5, (now - timedelta(hours=4)).isoformat(), now))
        self.assertTrue(synchronization_is_due(True, 5, (now - timedelta(hours=6)).isoformat(), now))


if __name__ == "__main__": unittest.main()

