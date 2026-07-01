import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import email_automation as app


class FakeAccounts:
    def smtp_configuration(self, email): return {"email": email, "password": "password", "smtp_host": "smtp.test", "smtp_port": 587, "encryption": "STARTTLS"}
    def connect_smtp(self, configuration, timeout=45): raise AssertionError("SMTP must not be reached")


class TrackingFailureTests(unittest.TestCase):
    def test_uuid_failure_marks_row_failed_and_does_not_send(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp); data = root / "data"; data.mkdir(); installed = root / "installed"; (installed / "config").mkdir(parents=True)
            (installed / "config" / "settings.json").write_text('{"data_dir":"' + str(data).replace("\\", "\\\\") + '","backup_enabled":false}', encoding="utf-8")
            wb = Workbook(); ws = wb.active
            ws.append(["Email", "Subject", "Body", "Sender_Email", "Status", "Result", "SentDate"])
            ws.append(["person@example.com", "Subject", "Body", "sender@example.com", "Pending", "", ""])
            wb.save(data / "mail_list.xlsx")
            with patch.object(app, "install_dir", return_value=installed), patch.object(app, "account_service", return_value=FakeAccounts()), patch.object(app, "domain_has_mx", return_value=True), patch.object(app.uuid, "uuid4", side_effect=RuntimeError("UUID generation failed")), patch.object(app.smtplib, "SMTP") as smtp:
                result = app.send_pending(1, wait_between=False)
                self.assertEqual(result["failed"], 1)
                smtp.assert_not_called()
            check = load_workbook(data / "mail_list.xlsx", data_only=True); sheet = check.active; headers = app.header_map(sheet)
            self.assertEqual(sheet.cell(2, headers["status"]).value, "Failed")
            self.assertIn("UUID generation failed", sheet.cell(2, headers["result"]).value)
            self.assertFalse(sheet.cell(2, headers["trackingid"]).value)
            check.close()

    def test_tracking_write_failure_prevents_smtp(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp); data = root / "data"; data.mkdir(); installed = root / "installed"; (installed / "config").mkdir(parents=True)
            (installed / "config" / "settings.json").write_text('{"data_dir":"' + str(data).replace("\\", "\\\\") + '","backup_enabled":false}', encoding="utf-8")
            wb = Workbook(); ws = wb.active
            ws.append(["Email", "Subject", "Body", "Sender_Email", "Status", "Result", "SentDate", "TrackingId"])
            ws.append(["person@example.com", "Subject", "Body", "sender@example.com", "Pending", "", "", ""])
            wb.save(data / "mail_list.xlsx")
            with patch.object(app, "install_dir", return_value=installed), patch.object(app, "account_service", return_value=FakeAccounts()), patch.object(app, "domain_has_mx", return_value=True), patch.object(app, "write_tracking_id", side_effect=RuntimeError("Failed to write TrackingId to mail_list.xlsx")), patch.object(app.smtplib, "SMTP") as smtp:
                result = app.send_pending(1, wait_between=False)
                self.assertEqual(result["failed"], 1)
                smtp.assert_not_called()
            check = load_workbook(data / "mail_list.xlsx", data_only=True); sheet = check.active; headers = app.header_map(sheet)
            self.assertEqual(sheet.cell(2, headers["status"]).value, "Failed")
            self.assertIn("Failed to write TrackingId", sheet.cell(2, headers["result"]).value)
            check.close()


if __name__ == "__main__": unittest.main()
