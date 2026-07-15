import json
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


TRACKING_COLUMNS = (
    "OpenCount",
    "FirstOpen",
    "LastOpen",
    "ClickCount",
    "FirstClick",
    "LastClick",
    "DownloadCount",
    "FirstDownload",
    "LastDownload",
    "ReplyCount",
    "FirstReply",
    "LastReply",
    "is_bounce",
    "bounce_time",
    "bounce_reason",
    "LastSynchronizeTime",
)
DOWNLOAD_COLUMNS = ("DownloadCount", "FirstDownload", "LastDownload")
BOUNCE_COLUMNS = ("is_bounce", "bounce_time", "bounce_reason")
MARK_SYNCHRONIZED_ENDPOINT = "/api/tracking/mark-synchronized"


class TrackingSynchronizationService:
    def __init__(self, base_url, excel_path, log_folder, http_get=None, http_post=None):
        self.base_url = base_url.rstrip("/")
        self.excel_path = Path(excel_path) if excel_path else None
        self.log_folder = Path(log_folder) if log_folder else None
        self.http_get = http_get or self._http_get
        self.http_post = http_post or self._http_post

    @staticmethod
    def _http_get(url):
        request = Request(url, headers={"Accept": "application/json", "User-Agent": "EmailAutomation/2"})
        with urlopen(request, timeout=60) as response:
            return json.loads(response.read().decode("utf-8"))

    @staticmethod
    def _http_post(url, payload):
        data = json.dumps(payload).encode("utf-8")
        request = Request(
            url,
            data=data,
            headers={
                "Accept": "application/json",
                "Content-Type": "application/json",
                "User-Agent": "EmailAutomation/2",
            },
            method="POST",
        )
        with urlopen(request, timeout=60) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else {}

    @staticmethod
    def _records(payload):
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            for key in ("records", "data", "results", "items"):
                if isinstance(payload.get(key), list):
                    return payload[key]
        raise ValueError("Tracking API returned an unsupported JSON format")

    @staticmethod
    def _normalized(record):
        return {str(key).replace("_", "").casefold(): value for key, value in record.items()}

    @classmethod
    def _maximum_updated_at(cls, records):
        maximum = None
        maximum_original = ""
        for record in records:
            if not isinstance(record, dict):
                continue
            value = cls._normalized(record).get("updatedat")
            if value in (None, ""):
                continue
            try:
                parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=timezone.utc)
                parsed = parsed.astimezone(timezone.utc)
            except (ValueError, TypeError):
                continue
            if maximum is None or parsed > maximum:
                maximum = parsed
                maximum_original = str(value)
        return maximum_original

    def _log(self, event, **values):
        return None

    @classmethod
    def _records_by_excel_path(cls, records):
        grouped = {}
        missing_excel_path = 0
        for record in records:
            if not isinstance(record, dict):
                continue
            values = cls._normalized(record)
            excel_path = str(values.get("excelfilepath") or "").strip()
            if not excel_path:
                missing_excel_path += 1
                continue
            grouped.setdefault(excel_path, []).append(record)
        return grouped, missing_excel_path

    def _mark_synchronized(self, tracking_id, last_synchronize_time):
        url = f"{self.base_url}{MARK_SYNCHRONIZED_ENDPOINT}"
        payload = {
            "tracking_id": tracking_id,
            "last_synchronize_time": last_synchronize_time,
        }
        return self.http_post(url, payload)

    def sync(self, last_sync_time="", full_synchronization_debug=False):
        started = time.monotonic()
        query = urlencode({"updated_after": last_sync_time}) if last_sync_time and not full_synchronization_debug else ""
        api_url = f"{self.base_url}/api/tracking/sync" + (f"?{query}" if query else "")
        self._log("Synchronization Started", api_url=api_url)
        try:
            payload = self.http_get(api_url)
            records = self._records(payload)
            maximum_updated_at = "" if full_synchronization_debug else self._maximum_updated_at(records)
            new_last_sync_time = last_sync_time if full_synchronization_debug else (maximum_updated_at or last_sync_time)
            grouped_records, records_missing_excel_path = self._records_by_excel_path(records)
            total_updated_rows = 0
            download_records_received = 0
            total_download_rows_updated = 0
            bounce_records_received = 0
            total_bounce_rows_updated = 0
            matched_tracking_ids = set()
            tracking_ids_not_found = set()
            missing_workbooks = set()
            workbooks_processed = 0
            synchronized_tracking_ids = []
            sync_completed_at = datetime.now(timezone.utc).isoformat()

            for excel_file_path, workbook_records in grouped_records.items():
                workbook_path = Path(excel_file_path)
                if not workbook_path.exists():
                    missing_workbooks.add(str(workbook_path))
                    self._log("Synchronization Warning", warning="Workbook missing", excel_file_path=str(workbook_path))
                    continue

                workbook = load_workbook(workbook_path)
                try:
                    worksheet = workbook.active
                    headers = {str(cell.value).strip().casefold(): index for index, cell in enumerate(worksheet[1], 1) if cell.value is not None}
                    tracking_column = headers.get("trackingid")
                    if not tracking_column:
                        raise ValueError(f"TrackingId column not found in {workbook_path}")

                    column_indexes = {}
                    columns_added = False
                    for name in TRACKING_COLUMNS:
                        key = name.casefold()
                        if key not in headers:
                            index = worksheet.max_column + 1
                            worksheet.cell(1, index, name)
                            headers[key] = index
                            columns_added = True
                        column_indexes[name] = headers[key]
                    has_workbook_bounce_data = any(any(column.replace("_", "").casefold() in self._normalized(record) for column in BOUNCE_COLUMNS) for record in workbook_records)
                    status_column = headers.get("status")
                    if has_workbook_bounce_data and not status_column:
                        status_column = worksheet.max_column + 1
                        worksheet.cell(1, status_column, "Status")
                        headers["status"] = status_column
                        columns_added = True

                    row_by_tracking_id = {}
                    for row in range(2, worksheet.max_row + 1):
                        value = worksheet.cell(row, tracking_column).value
                        if value not in (None, ""):
                            row_by_tracking_id[str(value).strip().casefold()] = row

                    workbook_updated_rows = set()
                    workbook_download_rows_updated = set()
                    workbook_bounce_rows_updated = set()
                    workbook_updated_tracking_ids = []

                    for record in workbook_records:
                        values = self._normalized(record)
                        has_download_data = any(name.replace("_", "").casefold() in values for name in DOWNLOAD_COLUMNS)
                        has_bounce_data = any(name.replace("_", "").casefold() in values for name in BOUNCE_COLUMNS)
                        if has_download_data:
                            download_records_received += 1
                        if has_bounce_data:
                            bounce_records_received += 1
                        tracking_id = values.get("trackingid")
                        normalized_tracking_id = str(tracking_id or "").strip().casefold()
                        row = row_by_tracking_id.get(normalized_tracking_id)
                        if not row:
                            if normalized_tracking_id:
                                tracking_ids_not_found.add(normalized_tracking_id)
                            continue

                        matched_tracking_ids.add(normalized_tracking_id)
                        for name in TRACKING_COLUMNS:
                            normalized_name = name.replace("_", "").casefold()
                            if normalized_name in values:
                                worksheet.cell(row, column_indexes[name], values[normalized_name])
                            elif name == "LastSynchronizeTime":
                                worksheet.cell(row, column_indexes[name], sync_completed_at)
                        if has_download_data:
                            workbook_download_rows_updated.add(row)
                        if has_bounce_data:
                            workbook_bounce_rows_updated.add(row)
                            if status_column:
                                worksheet.cell(row, status_column, "Bounce")
                            self._log("Excel Updated", excel_file_path=str(workbook_path), tracking_id=str(tracking_id or "").strip(), row=row, status="Bounce")
                        workbook_updated_rows.add(row)
                        workbook_updated_tracking_ids.append(str(tracking_id or "").strip())

                    if workbook_updated_rows or columns_added:
                        workbook.save(workbook_path)
                    workbooks_processed += 1
                    total_updated_rows += len(workbook_updated_rows)
                    total_download_rows_updated += len(workbook_download_rows_updated)
                    total_bounce_rows_updated += len(workbook_bounce_rows_updated)
                    synchronized_tracking_ids.extend(workbook_updated_tracking_ids)
                finally:
                    workbook.close()

            for tracking_id in synchronized_tracking_ids:
                self._mark_synchronized(tracking_id, sync_completed_at)

            elapsed = time.monotonic() - started
            result = {
                "records_downloaded": len(records),
                "rows_updated": total_updated_rows,
                "tracking_ids_matched": len(matched_tracking_ids),
                "tracking_ids_not_found": len(tracking_ids_not_found),
                "download_records_received": download_records_received,
                "download_rows_updated": total_download_rows_updated,
                "bounce_records_received": bounce_records_received,
                "bounce_rows_updated": total_bounce_rows_updated,
                "workbooks_processed": workbooks_processed,
                "missing_workbooks": len(missing_workbooks),
                "records_missing_excel_path": records_missing_excel_path,
                "execution_time": elapsed,
                "last_sync_time": new_last_sync_time,
                "api_url": api_url,
            }
            self._log(
                "Synchronization Finished",
                api_url=api_url,
                previous_last_sync=last_sync_time,
                maximum_updated_at_received=maximum_updated_at,
                new_last_sync=new_last_sync_time,
                records_downloaded=len(records),
                total_records_downloaded=len(records),
                total_tracking_ids_matched=len(matched_tracking_ids),
                rows_updated=total_updated_rows,
                total_excel_rows_updated=total_updated_rows,
                total_tracking_ids_not_found=len(tracking_ids_not_found),
                download_records_received=download_records_received,
                download_rows_updated=total_download_rows_updated,
                bounce_records_received=bounce_records_received,
                bounce_rows_updated=total_bounce_rows_updated,
                workbooks_processed=workbooks_processed,
                missing_workbooks=len(missing_workbooks),
                records_missing_excel_path=records_missing_excel_path,
                full_synchronization_debug=bool(full_synchronization_debug),
                execution_time_seconds=round(elapsed, 3),
            )
            return result
        except Exception as exc:
            elapsed = time.monotonic() - started
            self._log("Synchronization Error", api_url=api_url, error=str(exc), execution_time_seconds=round(elapsed, 3))
            raise


def synchronization_is_due(enabled, interval_hours, last_sync_time, now=None):
    if not enabled:
        return False
    if not last_sync_time:
        return True
    now = now or datetime.now(timezone.utc)
    try:
        previous = datetime.fromisoformat(last_sync_time.replace("Z", "+00:00"))
        if previous.tzinfo is None:
            previous = previous.replace(tzinfo=timezone.utc)
        return (now - previous).total_seconds() >= int(interval_hours) * 3600
    except (ValueError, TypeError):
        return True
