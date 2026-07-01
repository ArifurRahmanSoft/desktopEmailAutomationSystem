import email
import imaplib
import json
import re
from datetime import datetime
from email import policy
from email.header import decode_header, make_header
from pathlib import Path

from openpyxl import load_workbook


BOUNCE_COLUMNS = ("BounceStatus", "BounceReason", "BounceTime")
BOUNCE_SENDERS = ("mail delivery subsystem", "mailer-daemon", "mailer daemon", "postmaster")
BOUNCE_SUBJECT_WORDS = ("undeliver", "delivery status notification", "delivery failure", "mail delivery failed", "returned mail")
EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)


def decode_text(value):
    try: return str(make_header(decode_header(value or "")))
    except Exception: return str(value or "")


def message_text(message):
    values = []
    for part in message.walk():
        if part.get_content_type() == "text/plain":
            try: values.append(part.get_content())
            except Exception:
                payload = part.get_payload(decode=True) or b""
                values.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))
    return "\n".join(values)


def is_bounce_message(message):
    sender = decode_text(message.get("From")).casefold()
    subject = decode_text(message.get("Subject")).casefold()
    return any(value in sender for value in BOUNCE_SENDERS) or any(value in subject for value in BOUNCE_SUBJECT_WORDS)


def extract_bounce_details(message):
    recipient = ""
    reason = ""
    for part in message.walk():
        if part.get_content_type() != "message/delivery-status":
            continue
        payload = part.get_payload()
        blocks = payload if isinstance(payload, list) else [part]
        for block in blocks:
            for header in ("Final-Recipient", "Original-Recipient", "X-Failed-Recipients"):
                value = block.get(header)
                if value and not recipient:
                    matches = EMAIL_PATTERN.findall(str(value))
                    recipient = matches[0] if matches else ""
            diagnostic = block.get("Diagnostic-Code")
            if diagnostic and not reason:
                reason = str(diagnostic).split(";", 1)[-1].strip()
            if not reason and block.get("Status"):
                reason = f"Status {block.get('Status')}"
    text = message_text(message)
    if not recipient:
        patterns = (
            r"(?:Final-Recipient|Original-Recipient):\s*(?:rfc822;)?\s*([^\s<>;,]+@[^\s<>;,]+)",
            r"X-Failed-Recipients:\s*([^\s<>;,]+@[^\s<>;,]+)",
            r"(?:delivery to|recipient address|address)\s*[<:]?\s*([^\s<>;,]+@[^\s<>;,]+)",
        )
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match: recipient = match.group(1).strip(".>,; "); break
    if not reason:
        match = re.search(r"Diagnostic-Code:\s*(?:smtp;)?\s*(.+)", text, re.IGNORECASE)
        reason = match.group(1).strip() if match else decode_text(message.get("Subject")) or "Delivery failed"
    return recipient.lower(), reason[:1000]


class BounceTrackingService:
    def __init__(self, excel_path, log_folder, imap_ssl_factory=None, imap_factory=None):
        self.excel_path = Path(excel_path)
        self.log_folder = Path(log_folder)
        self.imap_ssl_factory = imap_ssl_factory or imaplib.IMAP4_SSL
        self.imap_factory = imap_factory or imaplib.IMAP4

    def _connect(self, configuration):
        encryption = configuration.get("imap_encryption", "SSL/TLS")
        factory = self.imap_ssl_factory if encryption == "SSL/TLS" else self.imap_factory
        client = factory(configuration["imap_host"], int(configuration["imap_port"]))
        if encryption == "STARTTLS": client.starttls()
        client.login(configuration["email"], configuration["password"])
        return client

    def test_connection(self, configuration):
        client = self._connect(configuration)
        try: client.noop()
        finally:
            try: client.logout()
            except Exception: pass
        return True

    def _log(self, **values):
        self.log_folder.mkdir(parents=True, exist_ok=True)
        path = self.log_folder / f"bounce-{datetime.now():%Y-%m-%d}.log"
        with path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps({"timestamp": datetime.now().isoformat(), **values}, ensure_ascii=False) + "\n")

    def check_account(self, configuration):
        client = self._connect(configuration)
        detected = matched = 0
        seen_message_ids = set()
        try:
            status, _ = client.select("INBOX")
            if status != "OK": raise RuntimeError("Unable to select IMAP inbox")
            status, data = client.search(None, "UNSEEN")
            if status != "OK": raise RuntimeError("Unable to search unread IMAP messages")
            for message_number in (data[0].split() if data and data[0] else []):
                status, payload = client.fetch(message_number, "(RFC822)")
                if status != "OK": continue
                raw = next((item[1] for item in payload if isinstance(item, tuple) and len(item) > 1), None)
                if not raw: continue
                message = email.message_from_bytes(raw, policy=policy.default)
                if not is_bounce_message(message): continue
                message_id = str(message.get("Message-ID") or message_number)
                if message_id in seen_message_ids:
                    client.store(message_number, "+FLAGS", "\\Seen"); continue
                seen_message_ids.add(message_id)
                recipient, reason = extract_bounce_details(message)
                detected += 1
                row = self.update_excel(recipient, reason) if recipient else None
                if row: matched += 1
                self._log(account=configuration["email"], recipient=recipient, reason=reason, excel_row=row, matched=bool(row), message_id=message_id)
                client.store(message_number, "+FLAGS", "\\Seen")
            return {"detected": detected, "matched": matched}
        finally:
            try: client.logout()
            except Exception: pass

    def update_excel(self, recipient, reason):
        if not self.excel_path.exists(): raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        workbook = load_workbook(self.excel_path)
        worksheet = workbook.active
        headers = {str(cell.value).strip().casefold(): index for index, cell in enumerate(worksheet[1], 1) if cell.value is not None}
        if "email" not in headers:
            workbook.close(); raise ValueError("Email column not found in mail_list.xlsx")
        indexes = {}
        for name in BOUNCE_COLUMNS:
            key = name.casefold()
            if key not in headers:
                column = worksheet.max_column + 1; worksheet.cell(1, column, name); headers[key] = column
            indexes[name] = headers[key]
        matched_row = None
        for row in range(2, worksheet.max_row + 1):
            value = str(worksheet.cell(row, headers["email"]).value or "").strip().casefold()
            if value == recipient.casefold(): matched_row = row
        if matched_row:
            if "status" in headers:
                worksheet.cell(matched_row, headers["status"], "Failed")
            if "result" in headers:
                worksheet.cell(matched_row, headers["result"], reason)
            worksheet.cell(matched_row, indexes["BounceStatus"], "Yes")
            worksheet.cell(matched_row, indexes["BounceReason"], reason)
            worksheet.cell(matched_row, indexes["BounceTime"], datetime.now().strftime("%d-%b-%Y %H:%M:%S"))
            workbook.save(self.excel_path)
        workbook.close()
        return matched_row
